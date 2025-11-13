
# All the variables are imported from LP.py file
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import os
import sys
import warnings
import re
from openpyxl import utils
import warnings
import uuid
# warnings.filterwarnings("ignore", category=UserWarning)
warnings.simplefilter("ignore", category=UserWarning, lineno=329, append=False)
warnings.filterwarnings('ignore', message='The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*',
                       category=FutureWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

pd.set_option('future.no_silent_downcasting', True)
parent_dir = os.path.dirname(os.getcwd())
sys.path.insert(0, parent_dir)
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from LP import *
from inference_grades import predict_grades

def module_4(input_folder, output_folder, params):
    print("Модуль 4: Выставление грейдов.")

    folder_py = params['folder_past_year']
    already_fixed = params['after_fix']
    
    counter = 0
    found_files=[]
    if not already_fixed: # Первичная обработка
        for file in os.listdir(input_folder):
            # Check if the file is an Excel file
            if file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm'):
                counter+=1
                output_file = os.path.join(output_folder, file)
                file_path = os.path.join(input_folder, file)
                filename, _ = os.path.splitext(file)

                print(f"Processing file {counter}: {file}")
                # df = pd.read_excel(file_path, sheet_name='Total Data')
                # df.to_excel(output_file)
                cols = [company_name, dep_level_1, dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6,
                                    job_title]
                
                df = pd.read_excel(file_path, sheet_name='Total Data')
                companies = df[company_name].unique()

                if not isinstance(folder_py, str) or not os.path.exists(folder_py):
                    print(f"Папка {folder_py} с анкетами прошлого года не найдена")
                elif os.path.exists(folder_py):
                    for company in companies:
                        print(f"Ищем анкету с прошлого года для компании {company}")
                        found_files = check_if_past_year_exist(company, folder_py)
                        if found_files:
                            file_to_cmp = os.path.join(folder_py, found_files[0])
                            df_py = pd.read_excel(file_to_cmp, sheet_name=rem_data, header=6)
                            cols_to_copy = [grade]
                            # Заполняем данными с прошлого года
                            df = merge_by_cols(df, df_py, cols, cols_to_copy)
                            # df.to_excel('debug.xlsx')
            
                
                # Делим данные на заполненные и незаполненные
                unfilled = df.loc[df[grade].apply(lambda x: str(x).lower().strip() == 'nan') == True] #add subfunction
                filled = df[~df.index.isin(unfilled.index)]
                empty_count = unfilled.shape[0]

                print(f"Проставлено грейдов: {len(filled)}, отсутствует грейдов: {len(unfilled)}")

                filled_and_processed = process_filled(filled)

                df, unfilled_and_processed, count_past_year, count_model = process_unfilled(unfilled, df)
                df.to_excel(output_file, sheet_name='Total Data')
                
                process_output_file(filled_and_processed, unfilled_and_processed, cols, output_file)

                info = {
                    'Файлы с прошлого года': str(found_files) if found_files else 'Отсутствуют',
                    'Отсутствующих грейдов в файле':  empty_count, 
                    'Всего строк в файле': df.shape[0],
                    'Подтянуто из прошлого года': count_past_year,
                    'Проставлено нейросетью': count_model
                }

                add_info(info, output_file)
                print("-----------------------")
                print()
    else: # Аналитик проверил и исправил анкету
        for file in os.listdir(output_folder):
            if file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm'):
                counter+=1
                output_file = os.path.join(output_folder, file)
                file_path = os.path.join(output_folder, file)
                filename, _ = os.path.splitext(file)

                print(f"Processing file {counter}: {file}")

                map_prefill_to_sheet1(file_path, output_file, sheet_prefill='Prefill')
                df_final = map_prefill_to_sheet1(file_path, output_file, sheet_prefill='Model')

                output_file = os.path.join(output_folder, filename+'_final.parquet')
                df_final.to_parquet(output_file)
                print(f"Файл {output_file} сохранен.")
                print("-----------------------")
                print()


def map_prefill_to_sheet1(
    excel_file: str,
    output_path,
    sheet_prefill,
    match_cols=[company_name, dep_level_1, dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6, job_title],
    code_cols=[grade],
    sheet_target='Total Data'
):
    """
    Маппит значения кодов из листа Prefill на данные в листе Sheet1 по совпадению колонок.
    """

    # --- читаем оба листа ---
    df_prefill = pd.read_excel(excel_file, sheet_name=sheet_prefill)
    df_target = pd.read_excel(excel_file, sheet_name=sheet_target)

    if match_cols is None:
        match_cols = [col for col in df_prefill.columns if col not in code_cols]

    # приведение типов к строке для колонок совпадений
    for col in match_cols:
        if col in df_prefill.columns:
            df_prefill[col] = df_prefill[col].astype(str).fillna('')
        if col in df_target.columns:
            df_target[col] = df_target[col].astype(str).fillna('')

    if set(match_cols).issubset(df_prefill.columns) and set(match_cols).issubset(df_target.columns):
        df_merged = df_target.merge(
            df_prefill[match_cols + list(code_cols)],
            on=match_cols,
            how='left',
            suffixes=('', '_prefill')
        )

        # --- заменяем коды из Prefill, если там есть значения ---
        for col in code_cols:
            df_merged[col] = df_merged[f"{col}_prefill"].combine_first(df_merged[col])
            df_merged.drop(columns=f"{col}_prefill", inplace=True)
        
        folder, filename = os.path.split(output_path)
        name, ext = os.path.splitext(filename)

        # Добавляем _processed
        processed_filename = f"{name}_processed{ext}"

        # Собираем обратно
        processed_path = os.path.join(folder, processed_filename)

        # сохраняем результат
        if not os.path.exists(processed_path):
            with pd.ExcelWriter(processed_path, engine="openpyxl", mode="w") as writer:
                df_merged.to_excel(writer, sheet_name=sheet_target, index=False)
        else:
            with pd.ExcelWriter(processed_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_merged.to_excel(writer, sheet_name=sheet_target, index=False)

        print(f"На лист '{sheet_target}' подтянуты значения из листа '{sheet_prefill}' в файле {processed_path}")
    else:
        print("Не все колонки из match_cols найдены в обоих листах.")
    
    return df_merged
    
        
def add_info(info, output_file):
    info = pd.DataFrame(data=[info])
    book = load_workbook(output_file)

    ws3 = book.create_sheet(title='Info')

    # Записываем заголовки
    for col_idx, col_name in enumerate(info.columns, start=1):
        ws3.cell(row=1, column=col_idx, value=col_name)

    # Записываем строки
    for row in info.itertuples(index=False):
        excel_row = ws3.max_row + 1
        for col_idx, value in enumerate(row, start=1):
            ws3.cell(row=excel_row, column=col_idx, value=value)
    book.save(output_file)


def process_output_file(df1, df2, cols, output_file, sheet1_name='Prefill', sheet2_name='Model'):
    """
    Добавляет два датафрейма в существующий Excel-файл.
    В df1 подсвечивает красным строки, где past_year_check == False.
    В df2 подсвечивает красным строки, где function_confidence < 70.
    """

    # Оставляем только уникальные строки
    df1 = df1.drop_duplicates(subset=cols)
    df2 = df2.drop_duplicates(subset=cols)

    if 'grade_old' in df1.columns:
        # print(1)
        df1 = df1.loc[:, [company_name,'Сектор', function_code, subfunction_code, specialization_code, grade,
               'past_year_check', dep_level_1, dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6,
                                job_title, 'grade_old']]
    else:
        # print(2)
        df1 = df1.loc[:, [company_name,'Сектор', function_code, subfunction_code, specialization_code, grade,
                        dep_level_1,
        dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6,
                        job_title]]
    
    if len(df2.columns)>1:
        if 'prediction_confidence' in df2.columns:
            # print(3)
            df2 = df2.loc[:, [company_name, function_code, subfunction_code, specialization_code, grade,
                'prediction_confidence',
                    dep_level_1, dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6,
                                    job_title]]
        else:
            # print(4)
            df2 = df2.loc[:, [company_name, function_code, subfunction_code, specialization_code,
                    dep_level_1, dep_level_2, dep_level_3, dep_level_4, dep_level_5, dep_level_6,
                                    job_title]]

    book = load_workbook(output_file)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Работа с df1
    ws1 = book.create_sheet(title=sheet1_name)
    for col_idx, col_name in enumerate(df1.columns, start=1):
        ws1.cell(row=1, column=col_idx, value=col_name)

    for _, row in df1.iterrows():
        excel_row = ws1.max_row + 1
        highlight = row.get('past_year_check') is False

        for col_idx, value in enumerate(row, start=1):
            cell = ws1.cell(row=excel_row, column=col_idx, value=value)
            if highlight:
                cell.fill = red_fill

    # Работа с df2
    ws2 = book.create_sheet(title=sheet2_name)
    for col_idx, col_name in enumerate(df2.columns, start=1):
        ws2.cell(row=1, column=col_idx, value=col_name)

    for _, row in df2.iterrows():
        excel_row = ws2.max_row + 1
        try:
            s = str(row.get('prediction_confidence'))
            num = float(s.rstrip('%'))
            highlight = num < 0.7
        except:
            highlight = False

        for col_idx, value in enumerate(row, start=1):
            cell = ws2.cell(row=excel_row, column=col_idx, value=value)
            if highlight:
                cell.fill = red_fill

    book.save(output_file)
    print(f"Листы '{sheet1_name}' и '{sheet2_name}' добавлены в файл: {output_file}")

def check_unfilled_columns(df):
    """
    Проверяет, есть ли пустые значения в колонке DataFrame. True если пропусков нет.
    """
    col = function_code
    # Приводим всё к строке и убираем пробелы
    mask_empty = df[col].astype(str).str.strip().isin(['', 'nan', 'NaN', 'None'])
    if mask_empty.any():
        print(f"Колонка '{col}' не заполнена полностью — есть пустые значения.")
        return
    return True

def process_unfilled(df, df_orig):
    # Подтягиваем коды прошлых лет в оригинальный датасет
    count_past_year = 0
    count_model = 0
    preds = pd.DataFrame()

    if check_unfilled_columns(df):
        if 'grade_old' in df_orig.columns:
            df_orig[function_code].update(df['grade_old'])

        df_without_py = df_orig.loc[df_orig[grade].apply(lambda x: str(x).lower().strip() == 'nan') == True]
        count_model = df_without_py.shape[0]
        count_past_year = df.shape[0] - count_model
        # Там где прошлого года нет, проставляем нейронкой
        if count_model != 0:
            preds = predict_grades(df_without_py)
            preds = preds.loc[preds[company_name].apply(lambda x: str(x).lower().strip() == 'nan') == False]

    return df_orig, preds, count_past_year, count_model
    

def process_filled(df):
    """
    Сравнивает столбцы grade и 'grade_old' в датафрейме.
    Создаёт новый столбец 'past_year_check', где:
      - True, если значения совпадают,
      - True, если значение func_old == NaN,
      - False — если различаются.
    """
    df = df.copy()
    df["past_year_check"] = True

    if "grade_old" in df.columns:
        df["past_year_check"] = (
            (df[grade] == df["grade_old"]) |
            (df["grade_old"].isna())
        )
    return df


def merge_by_cols(df, df_py, cols, cols_to_copy):
    """
    Сравнивает строки df и df_py по списку колонок cols.
    Если значения совпадают — копирует значения колонок cols_to_copy из df_py в df в колонку grade_old.

    Параметры:
        df (pd.DataFrame): основной датафрейм, в который копируем данные
        df_py (pd.DataFrame): источник данных
        cols (list): список колонок для сравнения
        cols_to_copy (list): список колонок, которые нужно скопировать из df_py в df (ожидается одна колонка: ['grade'])

    Возвращает:
        pd.DataFrame: обновлённый df
    """

    # Проверим наличие нужных колонок
    missing_cols = [c for c in cols + cols_to_copy if c not in df_py.columns]
    if missing_cols:
        raise ValueError(f"Отсутствуют колонки в df_py: {missing_cols}")

    df = df.copy()
    df_py = df_py.copy()

    # Приведение типов к строке (чтобы избежать ValueError при merge)
    for c in cols:
        df[c] = df[c].astype(str).replace('nan', np.nan)
        df_py[c] = df_py[c].astype(str).replace('nan', np.nan)

    # Уберём дубликаты по ключевым колонкам в df_py
    df_py_unique = df_py.drop_duplicates(subset=cols, keep="first")

    # Выполним объединение
    df_merged = df.merge(
        df_py_unique[cols + cols_to_copy],
        on=cols,
        how="left",
        suffixes=("", "_py")
    )

    # Теперь переносим данные из df_py в grade_old
    old_col = cols_to_copy[0]  # например, "grade"
    py_col = f"{old_col}_py"

    if py_col in df_merged.columns:
        df_merged["grade_old"] = df_merged[py_col]
        df_merged.drop(columns=[py_col], inplace=True)
    else:
        df_merged["grade_old"] = np.nan

    return df_merged


# def _normalize_val(v):
#     """Нормализация для сравнения: на str, strip и lower (None/NaN -> '')"""
#     if pd.isna(v):
#         return ""
#     s = str(v).strip()
#     return s.lower()

def check_if_past_year_exist(company, folder_py):
    company_str = str(company).strip()
    found_files = []

    for filename in os.listdir(folder_py):
        if company_str.lower() in filename.lower():
            found_files.append(filename)
    
    if found_files:
        for f in found_files:
            print(f"Найдена анкета прошлого года: {f}")
    else:
        print("Не найдено анкет прошлого года.")
    return found_files