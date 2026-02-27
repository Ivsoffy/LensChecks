import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import sys
import warnings
import numpy as np

warnings.simplefilter("ignore", category=UserWarning, lineno=329, append=False)
warnings.filterwarnings(
    "ignore",
    message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*",
    category=FutureWarning,
)
warnings.simplefilter(action="ignore", category=FutureWarning)
warnings.simplefilter(action="ignore", category=pd.errors.SettingWithCopyWarning)

pd.set_option("future.no_silent_downcasting", True)
parent_dir = os.path.dirname(os.getcwd())
sys.path.insert(0, parent_dir)

sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from ..LP import *
from .pipeline import CodeModel


cols = [
    company_name,
    dep_level_1,
    dep_level_2,
    dep_level_3,
    dep_level_4,
    dep_level_5,
    dep_level_6,
    job_title,
]

def module_2(input_folder, output_folder, params):
    """
    Summary: Process input Excel files and produce outputs for module 2.
    Args:
        input_folder (str): Folder with input Excel files.
        output_folder (str): Folder for output Excel files.
        params (dict): Settings including 'folder_past_year' and 'after_fix'.
    Returns:
        None
    Raises:
        ValueError: If folders or params are invalid.
    """
    if not isinstance(params, dict):
        raise ValueError("Ошибка: параметры должны быть словарем.")
    if "folder_past_year" not in params or "after_fix" not in params:
        raise ValueError("Ошибка: отсутствуют параметры 'folder_past_year' или 'after_fix'.")
    if not os.path.isdir(input_folder):
        raise ValueError(f"Ошибка: входная папка не найдена: {input_folder}")
    if not os.path.isdir(output_folder):
        raise ValueError(f"Ошибка: выходная папка не найдена: {output_folder}")

    folder_py = params["folder_past_year"]
    already_fixed = params["after_fix"]

    found_files = []
    for file in os.listdir(input_folder):
        if _is_excel_file(file):
            output_file = os.path.join(output_folder, file)
            input_file = os.path.join(input_folder, file)

            try:
                df = pd.read_excel(input_file, sheet_name="Total Data", index_col=0)
            except:
                try:
                    df = pd.read_excel(input_file, sheet_name="Данные", header=6)
                    df_company = pd.read_excel(input_file, sheet_name=company_data, header=1)
                    df[company_name] = df_company.iloc[0, 3]
                    df[gi_company_name] = df_company.iloc[0, 3]
                    df[gi_sector] = df_company.iloc[1, 3]
                    df[gi_origin] = df_company.iloc[2, 3]
                    df[gi_headcount_cat] = df_company.iloc[3, 3]
                    df[gi_revenue_cat] = df_company.iloc[4, 3]
                    df[gi_contact_name] = df_company.iloc[5, 3]
                    df[gi_title] = df_company.iloc[6, 3]
                    df[gi_tel] = df_company.iloc[7, 3]
                    df[gi_email] = df_company.iloc[8, 3]
                except Exception as e:
                    print(f"Ошибка чтения файла '{input_file}': {e}")
                    continue
                

            if function_code not in df.columns:
                print(f"Ошибка: отсутствует колонка '{function_code}' в файле '{input_file}'.")
                continue

            if not already_fixed:
                print("Модуль 2: Проставление кодов функций.")
                print("Проставление кодов из анкеты прошлого года (если она найдена)")
                df = process_past_year(folder_py, df)

                unfilled = df.loc[
                    df[function_code]
                    .astype(str)
                    .str.lower()
                    .str.strip()
                    .isin(["nan", "none", "null", ""])
                ]
                filled = df[~df.index.isin(unfilled.index)]
                empty_count = unfilled.shape[0]

                filled_and_processed = process_filled(filled)
                df, unfilled_and_processed, count_past_year, count_model = process_unfilled(unfilled, df)
                df.to_excel(output_file, sheet_name="Total Data")

                process_output_file(filled_and_processed, unfilled_and_processed, cols, output_file)

                info = {
                    "past_year_files": str(found_files) if found_files else "No files",
                    "empty_count": empty_count,
                    "total_count": df.shape[0],
                    "count_past_year": count_past_year,
                    "count_model": count_model,
                }

                add_info(info, output_file)
            else:
                print("Модуль 2: Проставление проверенных кодов функций.")

                map_prefill_to_sheet1(input_file, output_file, sheet_prefill="Prefill")
                df, _ = map_prefill_to_sheet1(input_file, output_file, sheet_prefill="Model")
                df = check_the_result(df)
                
                df = df.loc[:, ~df.columns.str.startswith("Unnamed:")]
                df.to_excel(output_file, sheet_name='Total Data')
                print(f"Файл {output_file} сохранен.")
            print(f"--------- Обработка файла {file} окончена ---------")
        

def _is_excel_file(filename):
    """
    Summary: Check if a filename is an Excel file.
    Args:
        filename (str): File name or path.
    Returns:
        bool: True if the extension is an Excel type.
    Raises:
        None
    """
    if not isinstance(filename, str):
        return False
    return filename.lower().endswith((".xlsx", ".xls", ".xlsm"))


def _write_df_to_sheet(ws, df, highlight_fn=None, fill=None):
    """
    Summary: Write a dataframe to a worksheet with optional highlighting.
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Target worksheet.
        df (pd.DataFrame): Dataframe to write.
        highlight_fn (callable | None): Function that returns True for highlighted rows.
        fill (openpyxl.styles.PatternFill | None): Fill style for highlighted rows.
    Returns:
        None
    Raises:
        None
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for _, row in df.iterrows():
        excel_row = ws.max_row + 1
        highlight = False
        if highlight_fn is not None:
            try:
                highlight = bool(highlight_fn(row))
            except Exception:
                highlight = False
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if highlight and fill is not None:
                cell.fill = fill


def process_past_year(folder_py, df):
    """
    Summary: Merge past-year codes into the current dataset by company.
    Args:
        folder_py (str): Folder path with past-year files.
        df (pd.DataFrame): Current input dataframe.
    Returns:
        pd.DataFrame: Updated dataframe with past-year codes if found.
    Raises:
        ValueError: If required columns are missing.
    """
    if company_name not in df.columns:
        raise ValueError("Ошибка: отсутствует обязательная колонка Название компании.")

    if not isinstance(folder_py, str) or not os.path.exists(folder_py):
        print(f"Предупреждение: путь к папке прошлогодних файлов некорректен: {folder_py}")
        return df

    companies = df[company_name].unique()

    for company in companies:
        found_files = []
        try:
            found_files = check_if_past_year_exist(company, folder_py)
            if found_files:
                file_to_cmp = os.path.join(folder_py, found_files[0])
                df_py = pd.read_excel(file_to_cmp, sheet_name=rem_data, header=6, index_col=None)
                cols_to_copy = [
                    function_code,
                    subfunction_code,
                    specialization_code,
                    function,
                    subfunction,
                    specialization,
                ]
                df = merge_by_cols(df, df_py, cols, cols_to_copy)
        except Exception as e:
            file_name = found_files[0] if found_files else "неизвестный файл"
            print(f"Ошибка при обработке прошлогоднего файла '{file_name}': {e}")
    return df


def check_column_rules(df, col_name, allowed_values):
    """
    Summary: Validate a column against allowed values.
    Args:
        df (pd.DataFrame): Input dataframe.
        col_name (str): Column name to validate.
        allowed_values (Iterable): Allowed values for the column.
    Returns:
        pd.DataFrame: Dataframe with 'errors_not_allowed' flags.
    Raises:
        ValueError: If column or allowed values are missing.
    """
    if col_name not in df.columns:
        raise ValueError(f"Ошибка: отсутствует колонка '{col_name}'.")
    if allowed_values is None:
        raise ValueError("Ошибка: список допустимых значений не задан.")

    df = df.copy()
    df[col_name] = df[col_name].astype(str).str.strip().str.upper()

    allowed_set = {str(v).strip().upper() for v in allowed_values}
    mask_allowed = df[col_name].isin(allowed_set)
    if col_name == specialization_code:
        mask_allowed = mask_allowed | (df[col_name] == "")

    if "errors_not_allowed" not in df.columns:
        df["errors_not_allowed"] = False
    df.loc[~mask_allowed, "errors_not_allowed"] = True

    return df


def check_the_result(df):
    """
    Summary: Validate codes against the SDF reference file and fill names.
    Args:
        df (pd.DataFrame): Input dataframe.
    Returns:
        pd.DataFrame: Validated dataframe with error flags and names filled.
    Raises:
        FileNotFoundError: If the SDF file is missing.
        ValueError: If required columns are missing.
    """
    sdf_path = "src/module_2/SDF.xlsx"
    if not os.path.exists(sdf_path):
        raise FileNotFoundError(f"Ошибка: файл SDF не найден: {sdf_path}")

    sdf = pd.read_excel(sdf_path, sheet_name="Каталог функций", header=4)
    allowed_funcs = sdf[function_code]
    allowed_subfuncs = sdf[subfunction_code]
    allowed_specs = sdf[specialization_code]

    df = check_column_rules(df, function_code, allowed_funcs)
    df = check_column_rules(df, subfunction_code, allowed_subfuncs)
    df = check_column_rules(df, specialization_code, allowed_specs)

    func = df[function_code].astype(str).str.strip()
    subfunc = df[subfunction_code].astype(str).str.strip()
    spec = df[specialization_code].astype(str).str.strip()

    df["errors_subfunc"] = func != subfunc.str[:2]
    df["errors_spec"] = ~((subfunc == spec.str[:3]) | (spec == "NAN"))

    df = fill_function_name_from_sdf(df, sdf, col_name=function_code, new_col_name=function, col_sdf="Название функции")
    df = fill_function_name_from_sdf(df, sdf, col_name=subfunction_code, new_col_name=subfunction, col_sdf="Название подфункции")
    df = fill_function_name_from_sdf(df, sdf, col_name=specialization_code, new_col_name=specialization, col_sdf="Специализация")
    return df


def fill_function_name_from_sdf(
    df,
    sdf,
    col_name,
    new_col_name,
    col_sdf,
):
    """
    Summary: Fill a name column from SDF based on a code column.
    Args:
        df (pd.DataFrame): Input dataframe.
        sdf (pd.DataFrame): Reference dataframe.
        col_name (str): Code column to map.
        new_col_name (str): Output name column.
        col_sdf (str): SDF column that contains the name.
    Returns:
        pd.DataFrame: Dataframe with the mapped name column.
    Raises:
        ValueError: If required columns are missing.
    """
    df = df.copy()

    missing_cols = [c for c in [col_name] if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Ошибка: отсутствуют колонки в данных: {missing_cols}")
    missing_sdf = [c for c in [col_name, col_sdf] if c not in sdf.columns]
    if missing_sdf:
        raise ValueError(f"Ошибка: отсутствуют колонки в SDF: {missing_sdf}")

    mapping = (
        sdf[[col_name, col_sdf]]
        .dropna()
        .drop_duplicates(subset=[col_name])
        .set_index(col_name)[col_sdf]
        .astype(str)
    )

    codes = df[col_name].astype(str).str.strip().str.upper()
    mapped = codes.map(mapping)

    df[new_col_name] = mapped

    return df


def map_prefill_to_sheet1(
    excel_file: str,
    output_path,
    sheet_prefill,
    match_cols=[
        company_name,
        dep_level_1,
        dep_level_2,
        dep_level_3,
        dep_level_4,
        dep_level_5,
        dep_level_6,
        job_title,
    ],
    code_cols=(function_code, subfunction_code, specialization_code),
    sheet_target="Total Data",
):
    """
    Summary: Map codes from a Prefill sheet to the target sheet using match columns.
    Args:
        excel_file (str): Source Excel file path.
        output_path (str): Output Excel file path.
        sheet_prefill (str): Prefill sheet name.
        match_cols (list): Columns to match records.
        code_cols (tuple): Code columns to copy.
        sheet_target (str): Target sheet name.
    Returns:
        tuple[pd.DataFrame, str]: Merged dataframe and processed output path.
    Raises:
        ValueError: If required inputs are invalid.
    """
    df_merged = pd.DataFrame()
    processed_path = output_path

    if not isinstance(excel_file, str) or not excel_file:
        raise ValueError("Ошибка: путь к исходному файлу не задан.")
    if not isinstance(output_path, str) or not output_path:
        raise ValueError("Ошибка: путь для выходного файла не задан.")

    try:
        try:
            df_prefill = pd.read_excel(excel_file, sheet_name=sheet_prefill)
            df_target = pd.read_excel(excel_file, sheet_name=sheet_target)
        except FileNotFoundError:
            print(f"Ошибка: файл не найден: '{excel_file}'.")
            return df_merged, output_path
        except ValueError as e:
            print(f"Ошибка чтения листа Excel: {e}")
            return df_merged, output_path
        except Exception as e:
            print(f"Ошибка чтения Excel: {e}")
            return df_merged, output_path

        if df_prefill.empty:
            return df_target, output_path

        if match_cols is None:
            match_cols = [col for col in df_prefill.columns if col not in code_cols]

        for col in match_cols:
            if col in df_prefill.columns:
                df_prefill[col] = df_prefill[col].astype(str).fillna("")
            if col in df_target.columns:
                df_target[col] = df_target[col].astype(str).fillna("")

        if set(match_cols).issubset(df_prefill.columns) and set(match_cols).issubset(df_target.columns):
            try:
                df_merged = df_target.merge(
                    df_prefill[match_cols + list(code_cols)],
                    on=match_cols,
                    how="left",
                    suffixes=("", "_prefill"),
                )
            except KeyError as e:
                print(f"Ошибка: отсутствует колонка при слиянии: {e}")
                return df_merged, output_path
            except Exception as e:
                print(f"Ошибка при слиянии данных: {e}")
                return df_merged, output_path

            for col in code_cols:
                try:
                    df_merged[col] = df_merged[f"{col}_prefill"].combine_first(df_merged[col])
                    df_merged.drop(columns=f"{col}_prefill", inplace=True)
                except KeyError:
                    print(f"Ошибка: колонка '{col}' отсутствует в данных Prefill.")
                except Exception as e:
                    print(f"Ошибка объединения колонки '{col}': {e}")

            folder, filename = os.path.split(output_path)
            name, ext = os.path.splitext(filename)
            processed_filename = f"{name}_processed{ext}"
            processed_path = os.path.join(folder, processed_filename)
        else:
            print("Ошибка: match_cols отсутствуют в Prefill или Target.")
            return df_merged, output_path
    except Exception as e:
        print(f"Ошибка при обработке Prefill: {e}")
        return df_merged, output_path
    return df_merged, processed_path


def add_info(info, output_file):
    """
    Summary: Add an Info sheet with summary statistics to the output file.
    Args:
        info (dict): Summary data to write.
        output_file (str): Excel file path to update.
    Returns:
        None
    Raises:
        FileNotFoundError: If the output file is missing.
    """
    info = pd.DataFrame(data=[info])
    try:
        book = load_workbook(output_file)
    except FileNotFoundError:
        raise FileNotFoundError(f"Ошибка: выходной файл не найден: {output_file}")

    ws3 = book.create_sheet(title="Info")

    for col_idx, col_name in enumerate(info.columns, start=1):
        ws3.cell(row=1, column=col_idx, value=col_name)

    for row in info.itertuples(index=False):
        excel_row = ws3.max_row + 1
        for col_idx, value in enumerate(row, start=1):
            ws3.cell(row=excel_row, column=col_idx, value=value)
    book.save(output_file)


def process_output_file(df1, df2, cols, output_file, sheet1_name="Prefill", sheet2_name="Model"):
    """
    Summary: Write processed datasets to output Excel file with highlights.
    Args:
        df1 (pd.DataFrame): Dataframe for Prefill sheet.
        df2 (pd.DataFrame): Dataframe for Model sheet.
        cols (list): Columns used for deduplication.
        output_file (str): Output Excel file path.
        sheet1_name (str): Name of the Prefill sheet.
        sheet2_name (str): Name of the Model sheet.
    Returns:
        None
    Raises:
        FileNotFoundError: If the output file is missing.
        ValueError: If required columns are missing.
    """
    df1 = df1.drop_duplicates(subset=cols)
    df2 = df2.drop_duplicates(subset=cols)

    base_cols_1 = [
        company_name,
        'Сектор',
        function_code,
        subfunction_code,
        specialization_code,
        dep_level_1,
        dep_level_2,
        dep_level_3,
        dep_level_4,
        dep_level_5,
        dep_level_6,
        job_title,
    ]
    extra_cols_1 = ["past_year_check", "func_old", "subfunc_old", "spec_old"]
    df1_cols = base_cols_1 + extra_cols_1 if "func_old" in df1.columns else base_cols_1
    df1 = df1[[c for c in df1_cols if c in df1.columns]]

    base_cols_2 = [
        company_name,
        'Сектор',
        "predicted_code",
        dep_level_1,
        dep_level_2,
        dep_level_3,
        dep_level_4,
        dep_level_5,
        dep_level_6,
        job_title,
    ]
    conf_cols = ["description"]
    df2_cols = base_cols_2[:5] + conf_cols + base_cols_2[5:] if "description" in df2.columns else base_cols_2
    df2 = df2[[c for c in df2_cols if c in df2.columns]]

    if "predicted_code" in df2.columns:
        # raise ValueError("Ошибка: отсутствует колонка predicted_code.")

        pred_codes = df2["predicted_code"].astype(str)
        df2[function_code] = pred_codes.str[:2]
        df2[subfunction_code] = pred_codes.str[:3]
        df2[specialization_code] = pred_codes.apply(lambda x: x[:5] if "-" in x else "")

    try:
        book = load_workbook(output_file)
    except FileNotFoundError:
        print(f"Ошибка: выходной файл не найден: {output_file}")
        return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _past_year_mismatch(row):
        return row.get("past_year_check") is False

    def _low_confidence(row):
        s = str(row.get("function_confidence"))
        try:
            num = float(s.rstrip("%"))
            return num < 70
        except Exception:
            return False

    ws1 = book.create_sheet(title=sheet1_name)
    _write_df_to_sheet(ws1, df1, highlight_fn=_past_year_mismatch, fill=red_fill)

    ws2 = book.create_sheet(title=sheet2_name)
    _write_df_to_sheet(ws2, df2, highlight_fn=_low_confidence, fill=red_fill)

    book.save(output_file)


def process_unfilled(df, df_orig):
    """
    Summary: Run model predictions for rows with missing codes.
    Args:
        df (pd.DataFrame): Unfilled rows to predict.
        df_orig (pd.DataFrame): Original dataframe to update.
    Returns:
        tuple[pd.DataFrame, pd.DataFrame, int, int]: Updated df, predictions, past-year count, model count.
    Raises:
        ValueError: If required columns are missing.
    """
    count_past_year = 0
    preds = pd.DataFrame()

    if "func_old" in df_orig.columns:
        df_orig[function_code].update(df["func_old"])
        df_orig[subfunction_code].update(df["subfunc_old"])
        df_orig[specialization_code].update(df["spec_old"])

    if function_code not in df_orig.columns:
        raise ValueError("Ошибка: отсутствует колонка function_code.")

    df_without_py = df_orig.loc[
        df_orig[function_code].astype(str).str.lower().str.strip().eq("nan")
    ]
    count_model = df_without_py.shape[0]
    count_past_year = df.shape[0] - count_model
    if count_model != 0:
        print("Проставление кодов нейросетью")
        model = CodeModel()
        preds = model.predict(df_without_py, test=True)

    return df_orig, preds, count_past_year, count_model


def process_filled(df):
    """
    Summary: Mark rows that match past-year codes.
    Args:
        df (pd.DataFrame): Dataframe with past-year code columns.
    Returns:
        pd.DataFrame: Updated dataframe with past_year_check flags.
    Raises:
        None
    """
    df = df.copy()
    df["past_year_check"] = True

    if "func_old" in df.columns:
        df["past_year_check"] = (
            (df[function_code] == df["func_old"]) | (df["func_old"].isna())
        )
    return df


def merge_by_cols(df, df_py, cols, cols_to_copy):
    """
    Summary: Merge df with df_py using key columns and copy code columns.
    Args:
        df (pd.DataFrame): Base dataframe.
        df_py (pd.DataFrame): Past-year dataframe.
        cols (list): Key columns for matching.
        cols_to_copy (list): Columns to copy from df_py.
    Returns:
        pd.DataFrame: Merged dataframe.
    Raises:
        ValueError: If required columns are missing.
    """
    missing_cols_df = [c for c in cols if c not in df.columns]
    if missing_cols_df:
        raise ValueError(f"Ошибка: отсутствуют колонки в текущих данных: {missing_cols_df}.")

    missing_cols = [c for c in cols + cols_to_copy if c not in df_py.columns]
    if missing_cols:
        raise ValueError(f"Ошибка: отсутствуют колонки в прошлогодних данных: {missing_cols}.")

    for c in cols:
        df[c] = df[c].astype(str).replace("nan", np.nan)
        df_py[c] = df_py[c].astype(str).replace("nan", np.nan)

    df_py_unique = df_py.drop_duplicates(subset=cols, keep="first")

    df_merged = df.merge(
        df_py_unique[cols + cols_to_copy],
        on=cols,
        how="left",
        suffixes=("", "_py"),
    )

    func_cols = ["func_old", "subfunc_old", "spec_old"]

    for old_col, new_col in zip(cols_to_copy, func_cols):
        py_col = f"{old_col}_py"
        if py_col in df_merged.columns:
            df_merged[new_col] = df_merged[py_col]
            df_merged.drop(columns=[py_col], inplace=True)
        else:
            df_merged[new_col] = np.nan
    return df_merged


def check_if_past_year_exist(company, folder_py):
    """
    Summary: Find past-year files for a company in a folder.
    Args:
        company (str): Company name to match.
        folder_py (str): Folder path with past-year files.
    Returns:
        list[str]: Matching filenames.
    Raises:
        ValueError: If the folder path is invalid.
    """
    if not os.path.isdir(folder_py):
        raise ValueError(f"Ошибка: папка не найдена: {folder_py}")

    company_str = str(company).strip()
    found_files = []

    for filename in os.listdir(folder_py):
        if company_str.lower() in filename.lower():
            found_files.append(filename)

    return found_files
