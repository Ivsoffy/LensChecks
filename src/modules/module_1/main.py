# All the variables are imported from LP.py file
import difflib
import os
import re
import sys
import time
import warnings

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# warnings.filterwarnings("ignore", category=UserWarning)
warnings.simplefilter("ignore", category=UserWarning, lineno=329, append=False)
warnings.filterwarnings(
    "ignore",
    message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*",
    category=FutureWarning,
)
pd.set_option("future.no_silent_downcasting", True)
parent_dir = os.path.dirname(os.getcwd())
sys.path.insert(0, parent_dir)

sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from .. import LP  # noqa: E402
from ..utils import (  # noqa: E402
    convert_some_columns_to_numeric,
    convert_some_columns_to_str,
    has_errors,
    init_errors,
    is_empty_value,
    is_excel_file,
    main_checks,
    normalize_column_names,
    prepare_total_data,
    sanitize_workbook_package,
    write_df_with_template,
)


def get_valid_path(prompt):
    while True:
        path = input(prompt)
        path = path.replace("\\", "/")  # Replace backslashes with forward slashes
        if os.path.isdir(path):
            return path
        else:
            print("Invalid path. Please try again.")


def check_general_info(errors, df_company, lang, df):
    # Setting columns names to the russian version

    df.columns = LP.expected_columns_rus

    try:  # добавить проверку на выпадающий список
        df[LP.company_name] = df_company.iloc[0, 1]
        df[LP.gi_company_name] = df_company.iloc[0, 1]
        df[LP.gi_sector] = df_company.iloc[1, 1]
        df[LP.gi_origin] = df_company.iloc[2, 1]
        df[LP.gi_headcount_cat] = df_company.iloc[3, 1]
        df[LP.gi_revenue_cat] = df_company.iloc[4, 1]
        df[LP.gi_contact_name] = df_company.iloc[5, 1]
        df[LP.gi_title] = df_company.iloc[6, 1]
        df[LP.gi_tel] = df_company.iloc[7, 1]
        df[LP.gi_email] = df_company.iloc[8, 1]
    except Exception as e:
        print(e)

    for _, row in df_company.loc[0:3, ["Unnamed: 2", "Unnamed: 3"]].iterrows():
        field_name = str(row["Unnamed: 2"]).strip()  # Название поля
        value = row["Unnamed: 3"]  # Значение
        if is_empty_value(value):
            errors["info_errors"].append(f"Incorrect General Info: {field_name}")

    comp_name = df[LP.company_name][0]
    if not re.fullmatch(r"[A-Za-z0-9_]+", str(comp_name)):
        errors["info_errors"] += [f"Incorrect company name format: {comp_name}"]

    df["SDF Language"] = lang
    return errors, df


def region_normalization(errors, text: str, index: int, lang) -> str:
    not_missing = not pd.isna(text)
    region_values = list(set(LP.final_region.values()))

    normalized_to_original = {value.strip().lower(): value for value in region_values}
    normalized_text = str(text).strip().lower() if not_missing else ""
    matched_text = normalized_to_original.get(normalized_text)
    if not matched_text and not_missing:
        closest_match = difflib.get_close_matches(
            normalized_text, normalized_to_original.keys(), n=1, cutoff=0.75
        )
        if closest_match:
            matched_text = normalized_to_original[closest_match[0]]

    in_dict_values = matched_text is not None
    if in_dict_values:
        text = matched_text
    if not (not_missing and in_dict_values):
        errors["data_errors"] += [(LP.region, index)]
    return text


# Function to assign values based on a mapping
def translate_values(df, columns, translation_map):
    """
    Translate values in specified DataFrame column(s) using a provided mapping dictionary.

    Parameters:
    df: pandas DataFrame
    columns: str or list of str, column name(s) to translate
    translation_map: dict, mapping of original values to translated values

    Returns:
    pandas DataFrame with translated values
    """
    # Create a copy to avoid modifying the original DataFrame
    df_copy = df.copy()

    # Ensure columns is a list for uniform processing
    if isinstance(columns, str):
        columns = [columns]

    # Apply translation to each specified column
    for col in columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].map(translation_map).fillna(df_copy[col])
        else:
            print(f"Warning: Column '{col}' not found in DataFrame")

    return df_copy


def map_column_values(df, check_column, amend_column, mapping_dict):
    """
    Check values in one column and assign mapped values to another column.

    Parameters:
    df: pandas DataFrame
    check_column: str, name of the column to check values in
    amend_column: str, name of the column to assign mapped values to
    mapping_dict: dict, mapping of check_column values to amend_column values

    Returns:
    pandas DataFrame with amended values
    """
    # Create a copy to avoid modifying the original DataFrame
    df_copy = df.copy()

    # Check if both columns exist
    if check_column not in df_copy.columns:
        print(f"Warning: Check column '{check_column}' not found in DataFrame")
        return df_copy

    if amend_column not in df_copy.columns:
        print(f"Warning: Amend column '{amend_column}' not found in DataFrame")
        return df_copy

    # Map values from check_column to amend_column
    df_copy[amend_column] = (
        df_copy[check_column].map(mapping_dict).fillna(df_copy[amend_column])
    )

    return df_copy


def eng_to_rus(df):
    # Apply translations using the tranlsation function | Converting English version to Russian
    df = translate_values(
        df, [LP.expat, LP.sti_eligibility, LP.lti_eligibility], LP.yes_no_map
    )
    df = translate_values(df, LP.man_emp, LP.manager_spec_map)
    df = translate_values(df, LP.performance, LP.performance_map)
    df = translate_values(df, LP.gender_id, LP.gender_map)
    df = translate_values(df, LP.tenure, LP.tenure_map)
    df = translate_values(df, [LP.lti_prog_1, LP.lti_prog_2, LP.lti_prog_3], LP.lti_map)
    df = translate_values(df, LP.gi_sector, LP.sector_map)
    df = translate_values(df, LP.gi_origin, LP.origin_map)
    df = translate_values(df, LP.gi_revenue_cat, LP.revenue_map)

    return df


def _normalize_excel_label(value):
    return str(value).strip().lower() if value is not None else ""


def _autofit_columns(ws, min_width=12, max_width=80):
    for col_idx in range(1, ws.max_column + 1):
        column_cells = next(
            ws.iter_cols(
                min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
            )
        )
        values = [
            len(str(cell.value)) for cell in column_cells if cell.value is not None
        ]
        if not values:
            continue
        width = max(min_width, min(max_width, max(values) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_errors_to_excel(errors, input_path, output_path):
    """Add an `Errors` sheet and highlight invalid cells without Excel COM."""
    input_path = os.path.abspath(input_path)
    output_path = os.path.abspath(output_path)

    info_errors = errors.get("info_errors", [])
    data_errors = errors.get("data_errors", [])
    unique_data_errors = list(dict.fromkeys(col for col, _ in data_errors))
    row_count = max(len(info_errors), len(unique_data_errors), 1)

    wb = load_workbook(input_path, keep_links=False)
    try:
        if "Errors" in wb.sheetnames:
            del wb["Errors"]
        ws_err = wb.create_sheet("Errors", 0)

        thin_side = Side(style="thin", color="000000")
        header_fill = PatternFill("solid", fgColor="4472C4")
        error_fill = PatternFill("solid", fgColor="FFC000")
        header_font = Font(bold=True, color="FFFFFF")
        centered = Alignment(horizontal="center", vertical="center")
        wrapped = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        ws_err["A1"] = "info_errors"
        ws_err["B1"] = "data_errors"
        for cell in ws_err[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border

        for row_idx in range(row_count):
            ws_err.cell(
                row=row_idx + 2,
                column=1,
                value=info_errors[row_idx] if row_idx < len(info_errors) else None,
            )
            ws_err.cell(
                row=row_idx + 2,
                column=2,
                value=(
                    unique_data_errors[row_idx]
                    if row_idx < len(unique_data_errors)
                    else None
                ),
            )
            for col_idx in (1, 2):
                cell = ws_err.cell(row=row_idx + 2, column=col_idx)
                cell.alignment = wrapped
                cell.border = border

        _autofit_columns(ws_err)

        data_sheet = next(
            (
                wb[sheet_name]
                for sheet_name in wb.sheetnames
                if _normalize_excel_label(sheet_name)
                in {"\u0434\u0430\u043d\u043d\u044b\u0435", "salary data"}
            ),
            None,
        )
        if data_sheet is None:
            raise ValueError("Expected data sheet not found.")

        header_row = 7
        data_start_row = header_row + 1
        header_map = {
            _normalize_excel_label(
                data_sheet.cell(row=header_row, column=col_idx).value
            ): col_idx
            for col_idx in range(1, data_sheet.max_column + 1)
            if data_sheet.cell(row=header_row, column=col_idx).value is not None
        }

        for col_name, idx in data_errors:
            col_idx = header_map.get(_normalize_excel_label(col_name))
            if not col_idx:
                print(f"Column {col_name} not found.")
                continue

            excel_row = data_start_row + idx
            if 1 <= excel_row <= data_sheet.max_row:
                data_sheet.cell(row=excel_row, column=col_idx).fill = error_fill
            else:
                print(f"Row {excel_row} doesn't exist.")

        wb.save(output_path)
    finally:
        wb.close()

    sanitize_workbook_package(output_path)

    print(
        f"The 'Errors' sheet has been added, and the cells are highlighted. File: {output_path}"
    )


def add_regions(errors, df, lang):
    df[LP.region] = df[LP.region].where(
        ~df[LP.region].isna()
        & (df[LP.region].astype(str).str.strip() != "")
        & (df[LP.region].astype(str).str.strip() != "-"),
        df[LP.region_client_fill],
    )
    df[LP.region] = df[LP.region].astype(str).str.lower()
    if lang == "ENG":
        df = translate_values(df, LP.region, LP.region_match_map)
    df = translate_values(df, LP.region, LP.final_region)
    df[LP.region] = df.apply(
        lambda x: region_normalization(errors, x[LP.region], x.name, lang), axis=1
    )

    df[LP.macroregion] = np.nan
    df = map_column_values(df, LP.region, LP.macroregion, LP.region_to_macroregion_map)
    return errors, df


def errors_rus_to_eng(errors):
    for ind, error in enumerate(errors["data_errors"]):
        new_error_ind = LP.expected_columns_rus.index(error[0])
        new_error = (LP.expected_columns_eng[new_error_ind], error[1])
        errors["data_errors"][ind] = new_error
    return errors


def check_and_process_data(errors, df, lang, params):
    df = convert_some_columns_to_numeric(df)
    df = convert_some_columns_to_str(df)

    if lang == "ENG":
        df = eng_to_rus(df)

    errors, df = add_regions(errors, df, lang)
    errors, df = main_checks(errors, df)

    if lang == "ENG":
        errors = errors_rus_to_eng(errors)

    return errors, df


def _print_unprocessed_summary(unprocessed_files):
    """Print concise validation summary for files with errors."""
    if not unprocessed_files:
        print("\nTest completed!")
        return

    print("=" * 20 + " WARNING! " + "=" * 20)
    print("List of unprocessed files:")
    for file_name, issue in unprocessed_files.items():
        data_err = [col for col, _ in issue.get("data_errors", [])]
        unique_data_err = list(dict.fromkeys(data_err))
        print(
            f"\nFile: {file_name}, Info errors: {issue['info_errors']}\nData errors: {unique_data_err}"
        )


def _save_unprocessed_files(unprocessed_files, output_folder):
    """Move files with validation errors to `unprocessed` and add an Errors sheet."""
    if not unprocessed_files:
        return

    unprocessed_folder = os.path.join(output_folder, "unprocessed")
    os.makedirs(unprocessed_folder, exist_ok=True)
    print(f"\nSaving {len(unprocessed_files)} files to 'unprocessed'...")

    for file_name, issue in unprocessed_files.items():
        source_path = os.path.join(output_folder, file_name)
        destination_path = os.path.join(unprocessed_folder, file_name)
        try:
            if not os.path.exists(source_path):
                continue
            if os.path.exists(destination_path):
                os.remove(destination_path)
            add_errors_to_excel(issue, source_path, destination_path)
            os.remove(source_path)
        except Exception as error:
            print(f"Failed to save file {file_name} to 'unprocessed': {str(error)}")


def _resolve_file_layout(sheet_names):
    """Detect language and required sheet names for one input workbook."""
    if "Salary Data" in sheet_names:
        return (
            "ENG",
            LP.rem_data_eng,
            LP.company_data_eng,
            [LP.company_name_eng, LP.job_title_eng],
        )
    return "RUS", LP.rem_data, LP.company_data, [LP.company_name, LP.job_title]


def _process_single_file(file_path, params):
    """Read and validate one workbook, returning processed dataframe and collected errors."""
    errors = init_errors()

    with pd.ExcelFile(file_path) as workbook:
        lang, rm_data_sheet, company_sheet, rows_to_drop = _resolve_file_layout(
            workbook.sheet_names
        )
        expected_columns = LP.set_expected_columns(lang)

        df = workbook.parse(rm_data_sheet, header=6)
        df = normalize_column_names(df)
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            errors["info_errors"].append(
                f"The following columns are missing from the Data: {missing_columns}"
            )
            return df, errors, lang

        df = df[expected_columns]
        df = prepare_total_data(df, rows_to_drop)
        df_company = workbook.parse(company_sheet, header=1).iloc[:, 2:]

    errors, df = check_general_info(errors, df_company, lang, df)
    errors, df = check_and_process_data(errors, df, lang, params)
    return df, errors, lang


def _save_single_db(result_frames, output_folder):
    """Save concatenated processed dataframes into one `result_db.xlsx` file."""
    if not result_frames:
        return

    result_df = pd.concat(result_frames)
    result_df = result_df.loc[:, ~result_df.columns.str.contains("^Unnamed")]
    file_output_path = os.path.join(output_folder, "result_db.xlsx")
    result_df.to_excel(file_output_path, sheet_name="Total Data")
    print(f"All files are combined into {output_folder}!")


def module_1(input_folder, output_folder, params=None):
    """Run module 1 validation and write processed/unprocessed outputs."""
    print("Module 1: Technical Validation.")
    process_start = time.time()

    unprocessed_files = file_processing(input_folder, output_folder, params=params)

    process_end = time.time()
    print(f"File processing took: {process_end - process_start}")
    _print_unprocessed_summary(unprocessed_files)
    _save_unprocessed_files(unprocessed_files, output_folder)
    return unprocessed_files


def file_processing(input_folder, output_folder, columns=None, params=None):
    """Process all supported Excel files from `input_folder` and return error mapping."""
    params = params or {}
    unprocessed_files = {}
    result_frames = []
    single_db = params.get("single_db", False)
    save_db_only_without_errors = params.get("save_db_only_without_errors", False)

    excel_files = [
        file_name
        for file_name in sorted(os.listdir(input_folder))
        if is_excel_file(file_name)
    ]
    for counter, file_name in enumerate(excel_files, start=1):
        print(f"Checking the file {counter}: {file_name}")
        file_path = os.path.join(input_folder, file_name)

        df, errors, lang = _process_single_file(file_path, params)
        file_has_errors = has_errors(errors)
        should_save = (not file_has_errors) or (not save_db_only_without_errors)

        if should_save:
            if single_db:
                result_frames.append(df)
            else:
                file_output_path = os.path.join(output_folder, file_name)
                df.to_excel(file_output_path, sheet_name="Total Data")
                print(f"File {file_name} is saved to {output_folder}!")

        if file_has_errors:
            base, ext = os.path.splitext(file_name)
            unprocessed_name = f"{base}_unprocessed_{ext}"
            file_output_path = os.path.join(output_folder, unprocessed_name)
            write_df_with_template(df, file_path, file_output_path, lang)
            unprocessed_files[unprocessed_name] = errors
        else:
            print("No errors were found in the file, congratulations!")

    if single_db:
        _save_single_db(result_frames, output_folder)

    return unprocessed_files
