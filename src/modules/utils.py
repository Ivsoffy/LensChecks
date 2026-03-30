import difflib
import os
import re
from copy import copy
from datetime import date, datetime

import numpy as np

# All the variables are imported from LP.py file
import pandas as pd
from modules import LP
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows

NUMERIC_COLUMNS_TO_COERCE = [
    LP.monthly_salary,
    LP.salary_rate,
    LP.number_monthly_salaries,
    LP.fact_sti,
    LP.fact_lti,
    LP.fact_lti_1,
    LP.fact_lti_2,
    LP.fact_lti_3,
    LP.target_lti_per,
    LP.additional_pay,
    LP.grade,
]

STRING_COLUMNS_TO_COERCE = [
    LP.man_emp,
    LP.gender_id,
    LP.sti_eligibility,
    LP.lti_eligibility,
    LP.expat,
    LP.performance,
    LP.function_code,
    LP.subfunction_code,
    LP.specialization_code,
]


def init_errors():
    """Create a standard error container used by validation modules."""
    return {
        "info_errors": [],
        "data_errors": [],
    }


def has_errors(errors):
    """Return True if at least one info or data error is present."""
    return bool(errors.get("info_errors") or errors.get("data_errors"))


def is_excel_file(file_name):
    """Check whether file name has a supported Excel extension."""
    return str(file_name).lower().endswith((".xlsx", ".xls", ".xlsm"))


def normalize_column_names(df):
    """Normalize DataFrame column names by trimming spaces and line breaks."""
    normalized_df = df.copy()
    normalized_df.columns = [
        re.sub(r"\s+", " ", str(col).replace("\n", " ").replace("\r", " ")).strip()
        for col in normalized_df.columns
    ]
    return normalized_df


def prepare_total_data(df, required_columns):
    """Drop rows where all required key columns are empty."""
    normalized_df = df.copy()
    existing_columns = [col for col in required_columns if col in normalized_df.columns]
    if not existing_columns:
        return normalized_df

    for column in existing_columns:
        normalized_df[column] = normalized_df[column].replace("", np.nan)
    normalized_df.dropna(subset=existing_columns, how="all", inplace=True)
    return normalized_df


def convert_some_columns_to_numeric(df, columns=None):
    """Convert configured numeric columns to numeric dtype with coercion."""
    numeric_columns = columns or NUMERIC_COLUMNS_TO_COERCE
    normalized_df = df.copy()
    for column in numeric_columns:
        if column not in normalized_df.columns:
            continue
        normalized_df[column] = (
            normalized_df[column]
            .astype(str)
            .str.replace(",", ".", regex=False)
            .str.replace("\xa0", "", regex=False)
        )
        normalized_df[column] = pd.to_numeric(normalized_df[column], errors="coerce")
    return normalized_df


def convert_some_columns_to_str(df, columns=None):
    """Convert configured columns to string dtype."""
    string_columns = columns or STRING_COLUMNS_TO_COERCE
    normalized_df = df.copy()
    for column in string_columns:
        if column not in normalized_df.columns:
            continue
        normalized_df[column] = normalized_df[column].astype(str)
    return normalized_df


def man_emp_normalization(errors, text: str, index) -> str:
    text = text.lower().strip()

    if not text or text == "nan" or text == "":
        return ""

    managers = [
        "руководитель",
        "руководители",
        "менеджер",
        "менеджеры",
        "manager",
        "managers",
    ]
    specialists = [
        "рабочий",
        "рабочие",
        "служащий",
        "служащие",
        "специалист",
        "специалисты",
        "specialist",
        "specialists",
    ]

    all_keywords = managers + specialists
    words = re.findall(r"\w+", text)

    for word in words:
        match = difflib.get_close_matches(word, all_keywords, n=1, cutoff=0.7)
        if match:
            if match[0] in managers:
                return "Руководитель"
            elif match[0] in specialists:
                return "Специалист"

    errors["data_errors"] += [(LP.man_emp, index)]
    return text


def expectation_normalization(text: str, index: int) -> str:
    valid = ["Соответствует ожиданиям", "Ниже ожиданий", "Выше ожиданий"]
    valid_eng = ["Meet expectations", "Below expectations", "Above expectations"]

    if not text or text.strip() == "" or text == "nan":
        return ""

    text = text.strip().lower()
    match = difflib.get_close_matches(text, [v.lower() for v in valid], n=1, cutoff=0.6)
    match_eng = difflib.get_close_matches(
        text, [v.lower() for v in valid_eng], n=1, cutoff=0.6
    )

    if match:
        for v in valid:
            if v.lower() == match[0]:
                return v
    elif match_eng:
        for ind in range(len(valid_eng)):
            if valid_eng[ind].lower() == match_eng[0]:
                return valid[ind]
    else:
        return ""


def level_normalization(value, index) -> str:
    """
    Преобразует значение в формат 'N-X' (где X от 1 до 20)
    """
    if value is not None:
        text = str(value).strip().upper()
        # Число из строки вроде 'N-3', 'n3', '3'
        match = re.search(r"(\d{1,2})", text)
        if match:
            num = int(match.group(1))
            if 1 <= num <= 20:
                return f"N-{num}"
    return ""


def number_monthly_salaries_normalization(errors, num, index):
    if pd.isna(num) or num == "":
        num = 12
    elif num < 12 or num > 15:
        errors["data_errors"] += [(LP.number_monthly_salaries, index)]
    return num


def expat_normalization(text: str, index: int) -> str:
    if is_empty_value(text):
        return text
    else:
        value = text.strip().lower()
        if value in ["да", "д", "yes", "y"]:
            return "Да"
        elif value in ["нет", "н", "no", "n"]:
            return "Нет"
        else:
            return ""


def gender_normalization(text: str, index: int) -> str:
    if text == "" or text == "nan":
        return ""

    text = text.lower().strip()

    woman = ["female", "женский", "жен", "f", "ж-й", "ж", "женщина"]
    man = ["male", "мужской", "муж", "m", "м-й", "м", "мужчина"]

    all_keywords = woman + man
    words = re.findall(r"\w+", text)

    for word in words:
        match = difflib.get_close_matches(word, all_keywords, n=1, cutoff=0.7)
        if match:
            if match[0] in woman:
                return "Ж"
            elif match[0] in man:
                return "М"

    return ""


def bod_normalization(errors, value, index, min_year=1936, max_year=2020):
    """
    Normalize year of birth to YYYY.
    Empty values are allowed. Invalid or unclear values add a data error.
    """
    bod = LP.bod

    if is_empty_value(value) or pd.isna(value):
        return np.nan

    # 1) datetime/date
    if isinstance(value, (pd.Timestamp, datetime, date, np.datetime64)):
        year = int(value.year)
        if min_year <= year <= max_year:
            return year
        errors["data_errors"] += [(bod, index)]
        return year

    # 2) числовые типы
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            year = int(value)
        except Exception:
            errors["data_errors"] += [(bod, index)]
            return value
        if min_year <= year <= max_year:
            return year
        errors["data_errors"] += [(bod, index)]
        return year

    # 3) строки
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            errors["data_errors"] += [(bod, index)]
            return value

        # если строка — это просто число
        if re.fullmatch(r"\d{4}", s):
            year = int(s)
            if min_year <= year <= max_year:
                return year
            errors["data_errors"] += [(bod, index)]
            return year

        m = re.search(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", s)
        if m:
            year = int(m.group(1))
            if min_year <= year <= max_year:
                return year
            errors["data_errors"] += [(bod, index)]
            return year

        # строка не число и не дата
        errors["data_errors"] += [(bod, index)]
        return value

    errors["data_errors"] += [(bod, index)]
    return value


def hired_date_normalization(errors, value, index, min_year=1940, max_year=2026):
    """
    Normalize hire date to dd.mm.yyyy.
    Empty values are allowed. Invalid or out-of-range values add a data error.
    """
    hired_date = LP.hired_date

    if is_empty_value(value) or pd.isna(value):
        return np.nan

    if isinstance(value, (pd.Timestamp, datetime, np.datetime64)):
        dt = pd.to_datetime(value, errors="coerce")
    else:
        dt = pd.to_datetime(str(value).strip(), dayfirst=True, errors="coerce")

    if pd.isna(dt):
        errors["data_errors"] += [(hired_date, index)]
        return value

    year = int(dt.year)
    if not (min_year <= year <= max_year):
        errors["data_errors"] += [(hired_date, index)]
        return value

    return dt.strftime("%d.%m.%Y")


def tenure_normalization(tenure_value, hired_value, index):
    """
    Normalize tenure values:
    - "меньше года"/"менее года" -> "Меньше года"
    - If hire date is present and < 1 year ago, set to "Меньше года"
    - Otherwise keep original value
    """

    if not is_empty_value(hired_value):
        dt = pd.to_datetime(str(hired_value).strip(), dayfirst=True, errors="coerce")
        if not pd.isna(dt):
            if (datetime.today().date() - dt.date()).days < 365:
                return "Меньше года"
            else:
                return ""
    if isinstance(tenure_value, str):
        s = tenure_value.strip().lower()
        if s in ("меньше года", "менее года"):
            return "Меньше года"
    return tenure_value


def grade_normalization(errors, value, index, min_grade=7, max_grade=30):
    """
    Normalize grade to numeric value in [min_grade, max_grade].
    Empty values are allowed. Invalid or out-of-range values add a data error.
    """
    grade = LP.grade

    if is_empty_value(value) or pd.isna(value):
        return np.nan

    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        errors["data_errors"] += [(grade, index)]
        return value

    if num % 1 != 0:
        errors["data_errors"] += [(grade, index)]
        return value

    num = int(num)
    if not (min_grade <= num <= max_grade):
        errors["data_errors"] += [(grade, index)]
        return value

    return num


def lti_eligibility_normalization(value, row, index, lti_cols):
    """
    Normalize LTI eligibility:
    - Must be "Да" or "Нет"
    - If empty and all LTI cols empty -> set "Нет" and print message
    - If any LTI cols have values -> set "Да"
    - "да"/"нет" normalized to "Да"/"Нет"
    """
    v = value
    if isinstance(v, str):
        s = v.strip().lower()
        if s == "да":
            return "Да"
        if s == "нет":
            return "Нет"

    has_lti_values = any((not is_empty_value(row[col])) for col in lti_cols)

    if has_lti_values:
        return "Да"

    if is_empty_value(v) or pd.isna(v):
        return "Нет"
    return v


def salary_rate_normalization(errors, num: int, index: int) -> str:
    if not num or pd.isna(num):
        num = 1
    elif num >= 1.5 or num <= 0:
        errors["data_errors"] += [(LP.salary_rate, index)]
    return num


def additional_pay_normalization(errors, value, index):
    if value < 0:
        errors["data_errors"] += [(LP.additional_pay, index)]

    return value


def eligibility_normalization(fact, target, value, index):
    if not pd.isna(value):
        value = value.strip().lower()

        if value in ["да", "д", "yes", "y"]:
            return "Да"
        else:
            return "Нет"

    else:
        if not (pd.isna(fact) or target == "nan"):
            return "Да"
        else:
            return "Нет"


def fact_sti_normalization(errors, eligibility, value, index):
    if eligibility == "Нет" and not pd.isna(value) and not value == 0:
        errors["data_errors"] += [(LP.fact_sti, index)]
    return value


def is_empty_value(x):
    """Возвращает True, если значение можно считать пустым."""
    if x is None:
        return True
    if isinstance(x, (float, np.floating)) and pd.isna(x):
        return True
    if isinstance(x, str):
        s = x.strip().lower()
        if s in ("", "nan", "none", "null", "n/a", "na", "-", "--"):
            return True
    if isinstance(x, (list, tuple, dict, set)) and len(x) == 0:
        return True
    return False


def normalize_employee_code(series, min_value=1, max_value=None):
    """
    Ensure employee_code has only unique integer values with no value/NaN.
    Invalid or duplicate values are replaced with random unique integers.
    """
    s = series.copy()
    numeric = pd.to_numeric(s, errors="coerce")

    invalid = numeric.isna() | s.apply(is_empty_value)

    # Invalidate non-integer numeric values (e.g., 12.5)
    fractional = (~numeric.isna()) & (numeric % 1 != 0)
    invalid = invalid | fractional

    # Normalize valid numeric values to int
    normalized = numeric.copy()
    normalized.loc[~numeric.isna()] = numeric.loc[~numeric.isna()].astype("int64")

    # Mark duplicates (except first occurrence) as invalid
    dup = normalized.duplicated(keep="first")
    invalid = invalid | dup

    existing = set(normalized[~invalid].astype(int).tolist())

    if max_value is None:
        if existing:
            max_value = max(existing) + len(s) + 1000
        else:
            max_value = min_value + len(s) + 1000

    rng = np.random.default_rng()
    for idx in s.index[invalid]:
        while True:
            candidate = int(rng.integers(min_value, max_value))
            if candidate not in existing:
                existing.add(candidate)
                s.loc[idx] = candidate
                break

    return pd.to_numeric(s, errors="coerce").astype("int64")


def target_sti_normalization(errors, value: str, index: int) -> str:
    if is_empty_value(value):
        return value

    s = str(value).strip()
    s_ns = s.replace(" ", "")

    # Разрешено: число, опционально десятичная часть
    if not re.fullmatch(r"\d+([.,]\d+)?", s_ns):
        errors["data_errors"] += [(LP.target_sti, index)]
        return value

    # Нормализация для дальнейших расчетов
    s_ns = s_ns.rstrip("%").replace(",", ".")

    if not (2 < (float(s_ns) * 100) < 300):
        errors["data_errors"] += [(LP.target_sti, index)]
        return value
    return s_ns


def to_num_or_zero(v):
    if is_empty_value(v) or pd.isna(v):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def lti_checks(errors, main_lti, lti_1, lti_2, lti_3, index, type_lti):
    main_val = to_num_or_zero(main_lti)
    sum_parts = to_num_or_zero(lti_1) + to_num_or_zero(lti_2) + to_num_or_zero(lti_3)

    if not ((main_val == sum_parts) | np.isnan(main_val)):
        errors["data_errors"] += [(type_lti, index)]
    return main_lti


def check_column_rules(df, col_name, allowed_values):
    """
    Summary: Validate a column against allowed values.
    Args:
        df (pd.DataFrame): Input dataframe.
        col_name (str): Column name to validate.
        allowed_values (Iterable): Allowed values for the column.
    Returns:
        pd.DataFrame: Dataframe with 'errors_not_allowed' flags.
    Raises:
        ValueError: If column or allowed values are missing.
    """
    if col_name not in df.columns:
        raise ValueError(f"Ошибка: отсутствует колонка '{col_name}'.")
    if allowed_values is None:
        raise ValueError("Ошибка: список допустимых значений не задан.")

    df = df.copy()
    normalized = df[col_name].astype(str).str.strip().str.upper()
    mask_empty = normalized.isin({"", "NAN", "NONE", "NULL"})
    allowed_set = {
        str(v).strip().upper() for v in allowed_values if not is_empty_value(v)
    }
    mask_allowed = normalized.isin(allowed_set)
    if col_name == LP.specialization_code:
        # Empty specialization code is valid, other code columns must not be empty.
        mask_invalid = (~mask_empty) & (~mask_allowed)
    else:
        mask_invalid = mask_empty | (~mask_allowed)

    flag_col = f"errors_not_allowed__{col_name}"
    df[flag_col] = mask_invalid
    if "errors_not_allowed" not in df.columns:
        df["errors_not_allowed"] = False
    df["errors_not_allowed"] = df["errors_not_allowed"] | mask_invalid

    # Keep invalid values normalized to empty to simplify downstream matching.
    df.loc[mask_invalid, col_name] = ""
    if col_name == LP.specialization_code:
        df.loc[mask_empty, col_name] = ""

    return df


def fill_function_name_from_sdf(
    df,
    sdf,
    col_name,
    new_col_name,
    col_sdf,
):
    """
    Summary: Fill a name column from SDF based on a code column.
    Args:
        df (pd.DataFrame): Input dataframe.
        sdf (pd.DataFrame): Reference dataframe.
        col_name (str): Code column to map.
        new_col_name (str): Output name column.
        col_sdf (str): SDF column that contains the name.
    Returns:
        pd.DataFrame: Dataframe with the mapped name column.
    Raises:
        ValueError: If required columns are missing.
    """
    df = df.copy()

    missing_cols = [c for c in [col_name] if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Ошибка: отсутствуют колонки в данных: {missing_cols}")
    missing_sdf = [c for c in [col_name, col_sdf] if c not in sdf.columns]
    if missing_sdf:
        raise ValueError(f"Ошибка: отсутствуют колонки в SDF: {missing_sdf}")

    mapping = (
        sdf[[col_name, col_sdf]]
        .dropna()
        .drop_duplicates(subset=[col_name])
        .set_index(col_name)[col_sdf]
        .astype(str)
    )

    codes = df[col_name].astype(str).str.strip().str.upper()
    mapped = codes.map(mapping)

    df[new_col_name] = mapped

    return df


def target_lti_normalization(errors, value, index, column):
    if is_empty_value(value):
        return value

    s = str(value).strip()
    s_ns = s.replace(" ", "").replace("\xa0", "")

    # Excel percent-formatted cells are often read by pandas as fractions (e.g. 23.88% -> 0.2388)
    if re.fullmatch(r"\d+([.,]\d+)?", s_ns):
        s_num = s_ns.replace(",", ".")
        num = float(s_num)
        if 0.01 <= num <= 3:
            return s_num

    errors["data_errors"] += [(column, index)]
    return value


def codes_not_correspond(errors, value, index, column):
    if value:
        errors["data_errors"] += [(column, index)]


def check_codes(errors, df):
    """
    Summary: Validate codes against the SDF reference file and fill names.
    Args:
        df (pd.DataFrame): Input dataframe.
    Returns:
        pd.DataFrame: Validated dataframe with error flags and names filled.
    Raises:
        FileNotFoundError: If the SDF file is missing.
        ValueError: If required columns are missing.
    """
    sdf_path = "modules/funcs_2026.parquet"
    if not os.path.exists(sdf_path):
        raise FileNotFoundError(f"Error: SDF file not found: {sdf_path}")

    sdf = pd.read_parquet(sdf_path)

    function_code = LP.function_code
    subfunction_code = LP.subfunction_code
    specialization_code = LP.specialization_code

    allowed_funcs = sdf[function_code]
    allowed_subfuncs = sdf[subfunction_code]
    allowed_specs = sdf[specialization_code]

    df = check_column_rules(df, function_code, allowed_funcs)
    df = check_column_rules(df, subfunction_code, allowed_subfuncs)
    df = check_column_rules(df, specialization_code, allowed_specs)

    row_invalid_any = pd.Series(False, index=df.index, dtype=bool)
    for flag_col in (
        f"errors_not_allowed__{function_code}",
        f"errors_not_allowed__{subfunction_code}",
        f"errors_not_allowed__{specialization_code}",
    ):
        if flag_col in df.columns:
            row_invalid_any = row_invalid_any | df[flag_col].astype(bool)

    if row_invalid_any.any():
        df.loc[
            row_invalid_any, [function_code, subfunction_code, specialization_code]
        ] = ""

    df = fill_function_name_from_sdf(
        df,
        sdf,
        col_name=function_code,
        new_col_name=LP.function,
        col_sdf="Название функции",
    )
    df = fill_function_name_from_sdf(
        df,
        sdf,
        col_name=subfunction_code,
        new_col_name=LP.subfunction,
        col_sdf="Название подфункции",
    )
    df = fill_function_name_from_sdf(
        df,
        sdf,
        col_name=specialization_code,
        new_col_name=LP.specialization,
        col_sdf="Специализация",
    )

    func = df[function_code].astype(str).str.strip()
    subfunc = df[subfunction_code].astype(str).str.strip()
    spec = df[specialization_code].astype(str).str.strip()

    code_is_empty = {"", "NAN", "NONE", "NULL"}
    func_is_empty = func.isin(code_is_empty)
    subfunc_is_empty = subfunc.isin(code_is_empty)
    df["errors_subfunc"] = (
        (~func_is_empty) & (~subfunc_is_empty) & (func != subfunc.str[:2])
    )
    spec_is_empty = spec.isin(["", "NAN", "NONE", "NULL"])
    df["errors_spec"] = ~((subfunc == spec.str[:3]) | spec_is_empty)

    df.apply(
        lambda x: codes_not_correspond(
            errors, x["errors_subfunc"], x.name, subfunction_code
        ),
        axis=1,
    )
    df.apply(
        lambda x: codes_not_correspond(
            errors, x["errors_spec"], x.name, specialization_code
        ),
        axis=1,
    )

    spec_flag_col = f"errors_not_allowed__{specialization_code}"
    if spec_flag_col in df.columns:
        df.apply(
            lambda x: codes_not_correspond(
                errors, x[spec_flag_col], x.name, specialization_code
            ),
            axis=1,
        )

    helper_cols = [
        f"errors_not_allowed__{function_code}",
        f"errors_not_allowed__{subfunction_code}",
        f"errors_not_allowed__{specialization_code}",
        "errors_not_allowed",
    ]
    drop_cols = [col for col in helper_cols if col in df.columns]
    if drop_cols:
        df.drop(columns=drop_cols, inplace=True)

    return df


def lti_prog_checks(errors, text, index, column):
    valid = [
        "Фантомные акции / Phantom stock",
        "Акции с ограничением / Restricted stock units (RSU)",
        "Restricted stock awards (RSA)",
        "Акции результативности / Performance stock units (PSU)",
        "Юнит результативности/долгосрочная премия / Performance unit/long-term cash",
        "Опцион на акции / Stock option",
        "Право на оценку акций / Stock appreciation rights (SAR)",
    ]

    if is_empty_value(text):
        return text

    text = str(text).strip().lower()
    match = difflib.get_close_matches(text, [v.lower() for v in valid], n=1, cutoff=0.6)

    if match:
        for v in valid:
            if v.lower() == match[0]:
                return v
    else:
        errors["data_errors"] += [(column, index)]
        return text


def lti_freq_checks(errors, value, index, column):
    valid = [0.25, 0.5, 1, 2, 3, 4]

    if is_empty_value(value):
        return value

    if int(value) not in valid:
        errors["data_errors"] += [(column, index)]
    return value


def region_coeff_normalization(errors, region, coef, index):
    if region in LP.regions_with_coeff and (is_empty_value(coef) or coef == 0):
        errors["data_errors"] += [(LP.region_coeff, index)]

    if region not in LP.regions_with_coeff and not (is_empty_value(coef) or coef == 0):
        errors["data_errors"] += [(LP.region_coeff, index)]


def main_checks(errors, df):
    # Подразделение 1 уровня
    df[LP.dep_level_1] = df[LP.dep_level_1].apply(
        lambda x: "-" if is_empty_value(x) else x
    )
    # Код сотрудника
    df[LP.employee_code] = normalize_employee_code(df[LP.employee_code])
    # Название должности
    df[LP.job_title] = df.apply(
        lambda x: (
            "-"
            if (not x[LP.job_title])
            or (str(x[LP.job_title]).strip() == "nan")
            or (str(x[LP.job_title]).strip() == "")
            else x[LP.job_title]
        ),
        axis=1,
    )
    # Руководитель/специалист
    df[LP.man_emp] = df.apply(
        lambda x: man_emp_normalization(errors, x[LP.man_emp], x.name), axis=1
    )
    # Оценка эффективности работы сотрудника
    df[LP.performance] = df.apply(
        lambda x: expectation_normalization(x[LP.performance], x.name), axis=1
    )
    # Уровень подчинения по отношению к Первому лицу компании
    df[LP.n_level] = df.apply(
        lambda x: level_normalization(x[LP.n_level], x.name), axis=1
    )
    # Пол
    df[LP.gender_id] = df.apply(
        lambda x: gender_normalization(x[LP.gender_id], x.name), axis=1
    )
    # Год рождения
    df[LP.bod] = df.apply(
        lambda x: bod_normalization(errors, x[LP.bod], x.name), axis=1
    )
    df[LP.bod] = df[LP.bod].astype(str)
    # Экспат
    df[LP.expat] = df.apply(lambda x: expat_normalization(x[LP.expat], x.name), axis=1)
    # Дата приема на работу
    df[LP.hired_date] = df.apply(
        lambda x: hired_date_normalization(errors, x[LP.hired_date], x.name), axis=1
    )
    # Стаж
    df[LP.tenure] = df.apply(
        lambda x: tenure_normalization(x[LP.tenure], x[LP.hired_date], x.name), axis=1
    )
    # Грейд
    df[LP.grade] = df.apply(
        lambda x: grade_normalization(errors, x[LP.grade], x.name), axis=1
    )
    # Коды функций, подфункций и специализаций
    df = check_codes(errors, df)
    # Право на участие в LTIP
    lti_cols = LP.lti_cols

    df[LP.lti_eligibility] = df.apply(
        lambda x: lti_eligibility_normalization(
            x[LP.lti_eligibility], x, x.name, lti_cols
        ),
        axis=1,
    )
    # Размер ставки
    df[LP.salary_rate] = df.apply(
        lambda x: salary_rate_normalization(errors, x[LP.salary_rate], x.name), axis=1
    )
    # Ежемесячный оклад
    df.dropna(subset=[LP.monthly_salary], inplace=True)
    # Районный коэффициент и северная надбавка в месяц
    df.apply(
        lambda x: region_coeff_normalization(
            errors, x[LP.region], x[LP.region_coeff], x.name
        ),
        axis=1,
    )
    # Число окладов в году
    df[LP.number_monthly_salaries] = df.apply(
        lambda x: number_monthly_salaries_normalization(
            errors, x[LP.number_monthly_salaries], x.name
        ),
        axis=1,
    )
    # Постоянные надбавки и доплаты (общая сумма за год)
    df[LP.additional_pay] = df.apply(
        lambda x: additional_pay_normalization(errors, x[LP.additional_pay], x.name),
        axis=1,
    )
    # Право на получение переменного вознаграждения
    df[LP.sti_eligibility] = df.apply(
        lambda x: eligibility_normalization(
            x[LP.fact_sti], x[LP.target_sti], x[LP.sti_eligibility], x.name
        ),
        axis=1,
    )
    # Фактическая премия
    df[LP.fact_sti] = df.apply(
        lambda x: fact_sti_normalization(
            errors, x[LP.sti_eligibility], x[LP.fact_sti], x.name
        ),
        axis=1,
    )
    # Целевая премия (%)
    df[LP.target_sti] = df.apply(
        lambda x: target_sti_normalization(errors, x[LP.target_sti], x.name), axis=1
    )
    # Фактическая стоимость всех предоставленных типов LTI за 1 год (AK)
    df[LP.fact_lti] = df.apply(
        lambda x: lti_checks(
            errors,
            x[LP.fact_lti],
            x[LP.fact_lti_1],
            x[LP.fact_lti_2],
            x[LP.fact_lti_3],
            x.name,
            LP.fact_lti,
        ),
        axis=1,
    )
    # Целевая стоимость вознаграждения 1 как % от базового оклада за 1 год
    df[LP.target_lti_1] = df.apply(
        lambda x: target_lti_normalization(
            errors, x[LP.target_lti_1], x.name, LP.target_lti_1
        ),
        axis=1,
    )
    df[LP.target_lti_2] = df.apply(
        lambda x: target_lti_normalization(
            errors, x[LP.target_lti_2], x.name, LP.target_lti_2
        ),
        axis=1,
    )
    df[LP.target_lti_3] = df.apply(
        lambda x: target_lti_normalization(
            errors, x[LP.target_lti_3], x.name, LP.target_lti_3
        ),
        axis=1,
    )
    # Целевая стоимость всех предоставленных типов LTI в % от базового оклада за 1 год
    df[LP.target_lti_per] = df.apply(
        lambda x: lti_checks(
            errors,
            x[LP.target_lti_per],
            x[LP.target_lti_1],
            x[LP.target_lti_2],
            x[LP.target_lti_3],
            x.name,
            LP.target_lti_per,
        ),
        axis=1,
    )
    # Тип программы
    prog_cols = [LP.lti_prog_1, LP.lti_prog_2, LP.lti_prog_3]
    for prog in prog_cols:
        df[prog] = df.apply(
            lambda x: lti_prog_checks(errors, x[prog], x.name, prog), axis=1
        )
    # Частота выплат
    freq_cols = [LP.lti_pay_freq_1, LP.lti_pay_freq_2, LP.lti_pay_freq_3]
    for freq in freq_cols:
        df[freq] = df.apply(
            lambda x: lti_freq_checks(errors, x[freq], x.name, freq), axis=1
        )

    return errors, df


def _norm(value):
    return str(value).strip()


def _get_sheet(wb, target_name):
    for s in wb.sheetnames:
        if s.strip() == target_name:
            return wb[s]
    raise ValueError(f'Sheet "{target_name}" not found')


def _build_header_map(ws, header_row):
    header_map = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        if isinstance(cell, MergedCell):
            continue
        if cell.value is None:
            continue
        header_map[_norm(cell.value)] = c
    return header_map


def _clear_values_below(ws, header_row):
    for row in ws.iter_rows(
        min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _write_common_columns(ws, df, header_row, header_map):
    for name, col_idx in header_map.items():
        cell = ws.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = name

    for r_idx, row in enumerate(
        df.itertuples(index=False, name=None), start=header_row + 1
    ):
        for name, value in zip(df.columns, row):
            col_idx = header_map.get(_norm(name))
            if not col_idx:
                continue
            cell = ws.cell(row=r_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = value


def _write_full(ws, df, header_row):
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=header_row
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = value


def _extend_styles(ws, header_row, last_data_row):
    if last_data_row <= ws.max_row:
        return

    style_src_row = header_row + 1 if (header_row + 1) <= ws.max_row else header_row
    for r in range(ws.max_row + 1, last_data_row + 1):
        for c in range(1, ws.max_column + 1):
            src = ws.cell(row=style_src_row, column=c)
            dst = ws.cell(row=r, column=c)
            if src.has_style:
                dst._style = copy(src._style)
            if src.hyperlink:
                dst._hyperlink = copy(src.hyperlink)
            if src.comment:
                dst.comment = copy(src.comment)


def write_df_with_template(
    df,
    template_path,
    out_path,
    lang,
    header_row=7,
    only_common_columns=True,
):
    df = df.copy()
    df = df.loc[:, LP.expected_columns_rus]
    df.columns = df.columns.map(str)
    if lang == "RUS":
        sheet_name = "Данные"
    else:
        sheet_name = "Salary Data"
        df.columns = LP.expected_columns_eng

    wb = load_workbook(template_path)
    ws = _get_sheet(wb, sheet_name)

    header_map = _build_header_map(ws, header_row)
    _clear_values_below(ws, header_row)

    if only_common_columns:
        common_cols = [c for c in df.columns if _norm(c) in header_map]
        df = df[common_cols]
        _write_common_columns(ws, df, header_row, header_map)
    else:
        _write_full(ws, df, header_row)

    last_data_row = header_row + df.shape[0]
    _extend_styles(ws, header_row, last_data_row)

    wb.save(out_path)
