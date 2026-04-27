# All the variables are imported from LP.py file
import os
import sys
import warnings

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# warnings.filterwarnings("ignore", category=UserWarning)
warnings.simplefilter("ignore", category=UserWarning, lineno=329, append=False)
warnings.filterwarnings(
    "ignore",
    message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*",
    category=FutureWarning,
)
warnings.simplefilter(action="ignore", category=FutureWarning)
warnings.simplefilter(action="ignore", category=pd.errors.SettingWithCopyWarning)

pd.set_option("future.no_silent_downcasting", True)
parent_dir = os.path.dirname(os.getcwd())
sys.path.insert(0, parent_dir)
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from modules.module_4.pipeline import GradePredictor  # noqa: E402

from .. import LP  # noqa: E402

DEFAULT_SUB_MODEL_PATH = "modules/module_4/models/model_1/grade_model_weights/model.cbm"
DEFAULT_LEAD_MODEL_PATH = (
    "modules/module_4/models/model_2/grade_model_weights/model.cbm"
)
cols = [
    LP.company_name,
    LP.dep_level_1,
    LP.dep_level_2,
    LP.dep_level_3,
    LP.dep_level_4,
    LP.dep_level_5,
    LP.dep_level_6,
    LP.job_title,
]


def _empty_grade_mask(series):
    normalized = series.astype(str).str.strip().str.lower()
    return series.isna() | normalized.isin({"", "nan", "none", "null"})


def _fill_missing_grades_from_old(df):
    df = df.copy()
    if "grade_old" not in df.columns or LP.grade not in df.columns:
        return df
    empty_mask = _empty_grade_mask(df[LP.grade])
    df.loc[empty_mask, LP.grade] = df.loc[empty_mask, "grade_old"]
    return df


def _is_excel_file(filename):
    """
    Summary: Check if a filename is an Excel file.
    Args:
        filename (str): File name or path.
    Returns:
        bool: True if the extension is an Excel type.
    Raises:
        None
    """
    if not isinstance(filename, str):
        return False
    return filename.lower().endswith((".xlsx", ".xls", ".xlsm"))


def process_past_year(folder_py, df):
    """
    Summary: Merge past-year grades into the current dataset by company.
    Args:
        folder_py (str): Folder path with past-year files.
        df (pd.DataFrame): Current input dataframe.
    Returns:
        pd.DataFrame: Updated dataframe with past-year grades if found.
    Raises:
        ValueError: If required columns are missing.
    """
    if not isinstance(folder_py, str) or not os.path.exists(folder_py):
        print(f"Warning: invalid path to last year's files folder: {folder_py}")
        return df, []

    companies = df[LP.company_name].unique()

    for company in companies:
        found_files = []
        try:
            found_files = check_if_past_year_exist(company, folder_py)
            if found_files:
                file_to_cmp = os.path.join(folder_py, found_files[0])
                df_py = pd.read_excel(
                    file_to_cmp, sheet_name=LP.rem_data, header=6, index_col=None
                )
                cols_to_copy = [LP.grade]
                df = merge_by_cols(df, df_py, cols, cols_to_copy)
        except Exception as e:
            file_name = found_files[0] if found_files else "unknown file"
            print(f"Error processing last year's file '{file_name}': {e}")
    return df, found_files


def add_comparison_with_median(df):
    cols = [
        LP.job_title,
        LP.region,
        LP.function_code,
        LP.subfunction_code,
        LP.specialization_code,
    ]
    df = df.copy()

    required_cols = cols + [LP.base_pay]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        df["median_group_id"] = pd.NA
        df["median"] = np.nan
        df["comparison_with_median"] = "Average"
        return df

    # Normalize pay values to numeric for median/ratio calculations.
    df["_base_pay_numeric"] = pd.to_numeric(
        df[LP.base_pay]
        .astype(str)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce",
    )

    grouped = df.groupby(cols, dropna=False, sort=False)
    df["median_group_id"] = (grouped.ngroup() + 1).astype("Int64")
    median_by_group = grouped["_base_pay_numeric"].transform("median")
    ratio_to_median = df["_base_pay_numeric"] / median_by_group
    df["median"] = median_by_group

    comparison_with_median = pd.Series("Average", index=df.index, dtype="object")
    comparison_with_median[ratio_to_median >= 2] = "More"
    comparison_with_median[ratio_to_median <= 0.5] = "Less"

    df["comparison_with_median"] = comparison_with_median
    df.drop(columns=["_base_pay_numeric"], inplace=True)
    return df


def module_4(input_folder, output_folder, params):
    folder_py = params["folder_past_year"]
    already_fixed = params["after_fix"]

    counter = 0
    found_files = []
    for file in os.listdir(input_folder):
        if _is_excel_file(file):
            output_file = os.path.join(output_folder, file)
            input_file = os.path.join(input_folder, file)

            print(f"Processing file {counter}: {file}")
            df = pd.read_excel(input_file, sheet_name="Total Data", index_col=0)

            if not already_fixed:
                print("Module 4: Assigning grades.")
                print("Assigning grades from last year's questionnaire (if found)")
                df["grade_old"] = np.nan
                initial_empty_mask = _empty_grade_mask(df[LP.grade])
                df, found_files = process_past_year(folder_py, df)
                df = _fill_missing_grades_from_old(df)
                df = add_comparison_with_median(df)

                remaining_empty_mask = _empty_grade_mask(df[LP.grade])
                unfilled = df.loc[remaining_empty_mask]
                filled = df.loc[~remaining_empty_mask]
                # empty_count = unfilled.shape[0]
                empty_count = initial_empty_mask.sum()
                count_past_year = int(
                    initial_empty_mask.sum() - remaining_empty_mask.sum()
                )
                count_model = int(remaining_empty_mask.sum())

                print(
                    f"Assigned grades: {len(filled)}, missing grades: {len(unfilled)}"
                )

                filled_and_processed = process_filled(filled)

                df, unfilled_and_processed = process_unfilled(unfilled, df)
                df.to_excel(output_file, sheet_name="Total Data")

                process_output_file(
                    filled_and_processed, unfilled_and_processed, output_file
                )

                info = {
                    "past_year_files": str(found_files) if found_files else "No files",
                    "empty_count": empty_count,
                    "total_count": df.shape[0],
                    "count_past_year": count_past_year,
                    "count_model": count_model,
                }

                add_info(info, output_file)
            else:
                print("Module 4: Assigning validated grades.")
                df_final = pd.read_excel(input_file, sheet_name="Total Data")
                df_final = map_prefill_to_sheet1(
                    input_file, sheet_prefill="Prefill", df_target=df_final
                )
                df_final = map_prefill_to_sheet1(
                    input_file, sheet_prefill="Model", df_target=df_final
                )

                _, filename = os.path.split(output_file)

                output_file = os.path.join(output_folder, filename)
                df_final = df_final.loc[:, ~df_final.columns.str.startswith("Unnamed:")]
                df_final.to_excel(output_file, sheet_name="Total Data")
                print(f"File {output_file} saved.")
            print(f"--------- Processing file {file} completed ---------")


def map_prefill_to_sheet1(
    excel_file: str,
    sheet_prefill,
    df_target=None,
    match_cols=[
        LP.company_name,
        LP.dep_level_1,
        LP.dep_level_2,
        LP.dep_level_3,
        LP.dep_level_4,
        LP.dep_level_5,
        LP.dep_level_6,
        LP.job_title,
        "comparison_with_median",
    ],
    code_cols=[LP.grade],
    sheet_target="Total Data",
):
    """
    Summary: Map code values from a Prefill/Model sheet to the target sheet by match columns.
    Args:
        excel_file (str): Source Excel file path.
        sheet_prefill (str): Name of the source sheet with code values.
        df_target (pd.DataFrame | None): Optional target dataframe. If None, loaded from sheet_target.
        match_cols (list): Columns used to match rows between sheets.
        code_cols (list): Columns to copy from source sheet.
        sheet_target (str): Name of the target sheet.
    Returns:
        pd.DataFrame: Target dataframe with mapped values.
    Raises:
        None
    """

    xls = pd.ExcelFile(excel_file)
    if df_target is None:
        df_target = pd.read_excel(excel_file, sheet_name=sheet_target)
    else:
        df_target = df_target.copy()

    if sheet_prefill not in xls.sheet_names:
        print(f"Sheet '{sheet_prefill}' not found in file {excel_file}.")
        return df_target

    df_prefill = pd.read_excel(excel_file, sheet_name=sheet_prefill)
    if df_prefill.empty:
        return df_target

    if match_cols is None:
        match_cols = [col for col in df_prefill.columns if col not in code_cols]

    available_code_cols = [col for col in code_cols if col in df_prefill.columns]
    if not available_code_cols:
        return df_target

    for col in match_cols:
        if col in df_prefill.columns:
            df_prefill[col] = df_prefill[col].fillna("").astype(str).str.strip()
        if col in df_target.columns:
            df_target[col] = df_target[col].fillna("").astype(str).str.strip()

    if not (
        set(match_cols).issubset(df_prefill.columns)
        and set(match_cols).issubset(df_target.columns)
    ):
        print("Not all columns from match_cols were found in both sheets.")
        return df_target

    df_prefill_unique = df_prefill.drop_duplicates(subset=match_cols, keep="first")
    df_merged = df_target.merge(
        df_prefill_unique[match_cols + available_code_cols],
        on=match_cols,
        how="left",
        suffixes=("", "_prefill"),
    )

    for col in available_code_cols:
        prefill_col = f"{col}_prefill"
        if prefill_col in df_merged.columns:
            if col in df_merged.columns:
                df_merged[col] = df_merged[prefill_col].combine_first(df_merged[col])
            else:
                df_merged[col] = df_merged[prefill_col]
            df_merged.drop(columns=prefill_col, inplace=True)

    print(f"Values from sheet '{sheet_prefill}' were mapped to sheet '{sheet_target}'.")
    return df_merged


def add_info(info, output_file):
    info = pd.DataFrame(data=[info])
    book = load_workbook(output_file)

    ws3 = book.create_sheet(title="Info")

    for col_idx, col_name in enumerate(info.columns, start=1):
        ws3.cell(row=1, column=col_idx, value=col_name)

    for row in info.itertuples(index=False):
        excel_row = ws3.max_row + 1
        for col_idx, value in enumerate(row, start=1):
            ws3.cell(row=excel_row, column=col_idx, value=value)
    book.save(output_file)


def process_output_file(
    df1, df2, output_file, sheet1_name="Prefill", sheet2_name="Model"
):
    """
    Summary: Write two dataframes to an existing Excel file.
    Args:
        df1 (pd.DataFrame): Dataframe for sheet1_name.
        df2 (pd.DataFrame): Dataframe for sheet2_name.
        cols (list): Key columns for deduplication.
        output_file (str): Output Excel file path.
        sheet1_name (str): Name of the first output sheet.
        sheet2_name (str): Name of the second output sheet.
    Returns:
        None
    Raises:
        None
    """

    cols = [
        LP.company_name,
        LP.gi_sector,
        LP.dep_level_1,
        LP.dep_level_2,
        LP.dep_level_3,
        LP.dep_level_4,
        LP.dep_level_5,
        LP.dep_level_6,
        LP.job_title,
        LP.function_code,
        LP.subfunction_code,
        LP.specialization_code,
        LP.grade,
        LP.base_pay,
        "comparison_with_median",
        "median",
        "median_group_id",
    ]

    prefill_cols = cols + ["grade_old"]
    model_cols = cols

    if not df1.empty:
        df1 = df1.loc[:, prefill_cols]
    if not df2.empty:
        df2 = df2.loc[:, model_cols]

    unique_cols = LP.unique_cols_for_grade

    df1 = df1.drop_duplicates(subset=unique_cols)
    df2 = df2.drop_duplicates(subset=unique_cols)

    with pd.ExcelWriter(
        output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df1.to_excel(writer, sheet_name=sheet1_name, index=False)
        df2.to_excel(writer, sheet_name=sheet2_name, index=False)

    print(
        f"Sheets '{sheet1_name}' and '{sheet2_name}' were added to file: {output_file}"
    )


def check_unfilled_columns(df):
    """
    Summary: Check whether the grade column has empty values.
    Args:
        df (pd.DataFrame): Input dataframe.
    Returns:
        bool | None: True if no empty values are found, otherwise None.
    Raises:
        None
    """
    col = LP.grade
    mask_empty = df[col].astype(str).str.strip().isin(["", "nan", "NaN", "None"])
    if mask_empty.any():
        print(f"Column '{col}' is not fully filled - contains empty values.")
        return
    return True


def process_unfilled(df, df_orig):
    preds = pd.DataFrame()

    df_orig = _fill_missing_grades_from_old(df_orig)
    df_without_py = df_orig.loc[_empty_grade_mask(df_orig[LP.grade])]
    count_model = df_without_py.shape[0]
    if count_model != 0:
        model = GradePredictor(
            sub_predictor_kwargs={"path_to_model": DEFAULT_SUB_MODEL_PATH},
            lead_predictor_kwargs={"path_to_model": DEFAULT_LEAD_MODEL_PATH},
        )
        preds = model.predict(df_without_py)
        preds = preds.loc[~preds[LP.company_name].isna()]

    return df_orig, preds


def process_filled(df):
    """
    Summary: Compare grade with grade_old and mark consistency in past_year_check.
    Args:
        df (pd.DataFrame): Dataframe with grade columns.
    Returns:
        pd.DataFrame: Updated dataframe with past_year_check.
    Raises:
        None
    """
    df = df.copy()
    df["past_year_check"] = True

    df["past_year_check"] = (df[LP.grade] == df["grade_old"]) | (df["grade_old"].isna())
    return df


def merge_by_cols(df, df_py, cols, cols_to_copy):
    """
    Summary: Merge current data with past-year data by key columns and copy selected values.
    Args:
        df (pd.DataFrame): Current dataframe to update.
        df_py (pd.DataFrame): Past-year dataframe used as source.
        cols (list): Key columns for matching rows.
        cols_to_copy (list): Columns to copy from df_py.
    Returns:
        pd.DataFrame: Merged dataframe with grade_old filled when available.
    Raises:
        ValueError: If required columns are missing in df_py.
    """

    missing_cols = [c for c in cols + cols_to_copy if c not in df_py.columns]
    if missing_cols:
        raise ValueError(f"Missing columns in df_py: {missing_cols}")

    df = df.copy()
    df_py = df_py.copy()

    for c in cols:
        df[c] = df[c].astype(str).replace("nan", np.nan)
        df_py[c] = df_py[c].astype(str).replace("nan", np.nan)

    df_py_unique = df_py.drop_duplicates(subset=cols, keep="first")

    df_merged = df.merge(
        df_py_unique[cols + cols_to_copy], on=cols, how="left", suffixes=("", "_py")
    )

    old_col = cols_to_copy[0]
    py_col = f"{old_col}_py"

    if py_col in df_merged.columns:
        df_merged["grade_old"] = df_merged[py_col]
        df_merged.drop(columns=[py_col], inplace=True)
    else:
        df_merged["grade_old"] = np.nan

    return df_merged


def check_if_past_year_exist(company, folder_py):
    company_str = str(company).strip()
    found_files = []

    for filename in os.listdir(folder_py):
        if company_str.lower() in filename.lower():
            found_files.append(filename)

    if found_files:
        for f in found_files:
            print(f"Found last year's questionnaire: {f}")
    else:
        print("No last year's questionnaires found.")
    return found_files
