import shutil
import warnings
from pathlib import Path

import pandas as pd
from modules.module_1.main import module_1
from modules.module_2.main import module_2
from modules.module_3.main import module_3
from modules.module_4.main import module_4
from modules.module_5.main import module_5
from openpyxl import Workbook
from openpyxl.styles import Alignment
from yaml import SafeLoader, load

warnings.filterwarnings("ignore", category=UserWarning)

GENERAL_INFO_LABEL_TO_COLUMN = {
    "Название компании на английском языке": [
        "Название компании на английском языке",
        "Название компании (заполняется автоматически)",
    ],
    "Сектор": ["Сектор"],
    "Тип компании": ["Тип компании"],
    "Общее количество сотрудников по состоянию на 1 мая 2026 года": [
        "Общее количество сотрудников по состоянию на 1 мая 2026 года"
    ],
    "Выручка за 2025 год, руб.": ["Выручка за 2025 год, руб."],
}
DATA_SHEET_NAME = "Данные"
GENERAL_INFO_SHEET_NAME = "Общая информация"
TOTAL_DATA_SHEET_NAME = "Total Data"
DATA_HEADER_ROW = 7
DATA_START_ROW = 8
COMPANY_NAME_COLUMN = "Название компании (заполняется автоматически)"
GENERAL_INFO_TITLE = "Общая информация"
GENERAL_INFO_START_ROW = 3
GENERAL_INFO_LABEL_COLUMN = 3
GENERAL_INFO_VALUE_COLUMN = 4
EXPORT_DATA_HEADERS = [
    "Название компании (заполняется автоматически)",
    "Подразделение 1 уровня",
    "Подразделение 2 уровня",
    "Подразделение 3 уровня",
    "Подразделение 4 уровня",
    "Подразделение 5 уровня",
    "Подразделение 6 уровня",
    "Название должности",
    "Код сотрудника",
    "Код руководителя сотрудника",
    "Руководитель / специалист",
    "Оценка эффективности работы сотрудника",
    "Уровень подчинения по отношению к Первому лицу компании",
    "Экспат",
    "Пол",
    "Год рождения",
    "Дата приема на работу",
    "Сотрудники, проработавшие в компании меньше 1 года",
    "Название города",
    "Регион/область (заполняется автоматически)",
    "Внутренний грейд компании",
    "Грейд / Уровень обзора",
    "Код функции",
    "Код подфункции",
    "Код специализации",
    "Название функции (заполняется автоматически)",
    "Название подфункции (заполняется автоматически)",
    "Название специализации (заполняется автоматически)",
    "Размер ставки",
    "Ежемесячный оклад",
    "Число окладов в году",
    "Постоянные надбавки и доплаты (общая сумма за год)",
    "Право на получение переменного вознаграждения",
    "Фактическая премия",
    "Целевая премия (%)",
    "Право на участие в Программе долгосрочного вознаграждения (LTIP)",
    "Фактическая стоимость всех предоставленных типов LTI за 1 год",
    "Целевая стоимость всех предоставленных типов LTI в % от базового оклада за 1 год",
    "Тип программы 1",
    "Фактическая стоимость вознаграждения 1 за 1 год",
    "Целевая стоимость вознаграждения 1 как % от базового оклада за 1 год",
    "Частота выплат 1",
    "Тип программы 2",
    "Фактическая стоимость вознаграждения 2 за 1 год",
    "Целевая стоимость вознаграждения 2 как % от базового оклада за 1 год",
    "Частота выплат 2",
    "Тип программы 3",
    "Фактическая стоимость вознаграждения 3 за 1 год",
    "Целевая стоимость вознаграждения 3 как % от базового оклада за 1 год",
    "Частота выплат 3",
    "Комментарии",
    "Годовой оклад (AP)",
    "Базовый оклад (BP)",
    "Краткосрочное фактическое переменное вознаграждение (VP)",
    "Целевая Премия (TI)",
    "Фактическое совокупное вознаграждение (TC)",
    "Целевое совокупное вознаграждение (TTC)",
    "Фактическое долгосрочное вознаграждение (LTIP)",
    "Целевое долгосрочное вознаграждение (TLTIP)",
    "Прямое совокупное вознаграждение (TDC)",
    "Целевое прямое совокупное вознаграждение (TTDC)",
    "Макрорегион",
]


def load_config(config_path="config.yml"):
    with open(config_path, "r") as f:
        return load(f, Loader=SafeLoader)


def find_input_folder(input_folder, module, already_fixed):
    if input_folder and Path(input_folder).is_dir():
        return input_folder

    if module > 1:
        if (
            module != 2
            and module != 4
            or ((module == 2 or module == 4) and not already_fixed)
        ):
            input = "modules/module_" + str(module - 1) + "/output"
            if Path(input).is_dir():
                return input
            input = "modules/module_" + str(module) + "/input"
            if Path(input).is_dir():
                return input
        else:
            input = "modules/module_" + str(module) + "/output/require_check"
            if Path(input).is_dir():
                return input
    elif module == 1:
        input = "modules/module_" + str(module) + "/input"
        if Path(input).is_dir():
            return input

    return None


def find_output_folder(module, already_fixed):
    if (module != 2 and module != 4) or (
        (module == 2 or module == 4) and already_fixed
    ):
        output_folder = Path(f"modules/module_{module}/output")
    else:
        output_folder = Path(f"modules/module_{module}/output/require_check")
    output_folder.mkdir(parents=True, exist_ok=True)
    return output_folder


def _sanitize_result_filename(name, fallback_stem):
    invalid_chars = '<>:"/\\|?*'
    cleaned = str(name or "").strip()
    cleaned = "".join("_" if char in invalid_chars else char for char in cleaned)
    cleaned = cleaned.rstrip(". ")
    return cleaned or fallback_stem


def _get_first_non_empty_value(df, columns):
    for column in columns:
        if column not in df.columns:
            continue

        series = df[column]
        non_empty = series[series.notna()]
        if not non_empty.empty:
            return non_empty.iloc[0]
    return None


def _prepare_company_dataframe(df, data_headers):
    prepared_df = pd.DataFrame(index=df.index)
    for header in data_headers:
        if header in df.columns:
            prepared_df[header] = df[header]
        else:
            prepared_df[header] = None
    return prepared_df


def _build_general_info_sheet(ws, company_df):
    general_info_values = {
        label: _get_first_non_empty_value(company_df, columns)
        for label, columns in GENERAL_INFO_LABEL_TO_COLUMN.items()
    }
    for row_offset, label in enumerate(GENERAL_INFO_LABEL_TO_COLUMN, start=0):
        row_idx = GENERAL_INFO_START_ROW + row_offset
        ws.cell(row_idx, GENERAL_INFO_LABEL_COLUMN).value = label
        ws.cell(row_idx, GENERAL_INFO_VALUE_COLUMN).value = general_info_values[label]


def _build_data_sheet(ws, company_df):
    body_alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = f"A{DATA_START_ROW}"

    for col_idx, header in enumerate(EXPORT_DATA_HEADERS, start=1):
        cell = ws.cell(DATA_HEADER_ROW, col_idx)
        cell.value = header

    company_df = _prepare_company_dataframe(company_df, EXPORT_DATA_HEADERS)
    for row_offset, row_values in enumerate(
        company_df.itertuples(index=False), start=0
    ):
        excel_row = DATA_START_ROW + row_offset
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(excel_row, col_idx)
            cell.value = None if pd.isna(value) else value
            cell.alignment = body_alignment


def _create_company_workbook(company_df):
    workbook = Workbook()
    general_info_ws = workbook.active
    general_info_ws.title = GENERAL_INFO_SHEET_NAME
    _build_general_info_sheet(general_info_ws, company_df)

    data_ws = workbook.create_sheet(DATA_SHEET_NAME)
    _build_data_sheet(data_ws, company_df)
    return workbook


def _export_total_data_file(source_file, destination_folder):
    try:
        df = pd.read_excel(source_file, sheet_name=TOTAL_DATA_SHEET_NAME)
    except ValueError:
        shutil.copy2(source_file, destination_folder / source_file.name)
        return

    if COMPANY_NAME_COLUMN not in df.columns:
        shutil.copy2(source_file, destination_folder / source_file.name)
        return

    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df[df[COMPANY_NAME_COLUMN].notna()].copy()
    if df.empty:
        return

    for company_name, company_df in df.groupby(COMPANY_NAME_COLUMN, sort=False):
        workbook = _create_company_workbook(company_df)
        base_name = _sanitize_result_filename(company_name, source_file.stem)
        workbook.save(destination_folder / f"{base_name}.xlsx")
        workbook.close()


def _copy_module_5_results(src, dst):
    for item in dst.iterdir():
        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()

    for path in src.iterdir():
        if not path.is_file():
            continue

        if path.suffix.lower() == ".xlsx":
            _export_total_data_file(path, dst)
        else:
            shutil.copy2(path, dst / path.name)


def cleaning_folders():
    src = Path("modules/module_5/output")
    dst = Path("modules/results")
    dst.mkdir(exist_ok=True)
    _copy_module_5_results(src, dst)

    base = Path("modules")
    archive = base / "archive"

    for i in range(1, 6):
        for name in ("input", "output"):
            src_dir = base / f"module_{i}" / name
            if src_dir.is_dir():
                dst_dir = archive / f"module_{i}" / name
                dst_dir.mkdir(parents=True, exist_ok=True)
                for item in src_dir.iterdir():
                    dst_item = dst_dir / item.name
                    if dst_item.exists():
                        if dst_item.is_dir():
                            shutil.rmtree(dst_item)
                        else:
                            dst_item.unlink()
                    shutil.move(str(item), str(dst_item))


def run_checks(config):
    modules = [module_1, module_2, module_3, module_4, module_5]

    params_for_each_module = [
        # module_1
        ["save_db_only_without_errors", "single_db"],
        # module_2
        ["folder_past_year", "after_fix"],
        # module_3
        [],
        # module_4
        ["folder_past_year", "after_fix"],
        # module_5
        ["save_db_only_without_errors", "single_db"],
    ]

    input_folder = config["input_folder"]
    module_num = config["module"] - 1

    input_folder = find_input_folder(
        input_folder, config["module"], config["after_fix"]
    )
    output_folder = find_output_folder(config["module"], config["after_fix"])
    if input_folder is None:
        print(f"Directory input_folder: {input_folder} not found!")
        if (config["module"] == 2) or (config["module"] == 4):
            print("Check parameter after_fix в config.yml!")
        return

    print("Directory to check: ", input_folder)
    print("Checked files will be saved in the directory: ", output_folder)

    param_names = params_for_each_module[module_num]
    params = {}
    for param in param_names:
        params[param] = config[param]
    module_result = modules[module_num](
        input_folder=input_folder, output_folder=output_folder, params=params
    )

    if (config["module"] == 5) and config["clean"]:
        cleaning_folders()

    if config["module"] == 1 or config["module"] == 5:
        return module_result


if __name__ == "__main__":
    config = load_config()
    run_checks(config)
